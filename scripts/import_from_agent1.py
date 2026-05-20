#!/usr/bin/env python3
import argparse
import json
import re
import time
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook

from browser_session import BASE_URL, get_auth, get_default_bill_model, ui_save_bill_template, ui_template_name_id_map


def is_ok(resp):
    return resp.get("code") == 200 or resp.get("success") is True


def invalidate_cache_entry(cache, key):
    if cache is not None:
        cache.pop(key, None)


def query_company_users(company_id, headers, cache=None, force_refresh=False):
    cache_key = f"company_users:{company_id}"
    if cache is not None and not force_refresh and cache_key in cache:
        return cache[cache_key]
    users = (
        requests.post(
            f"{BASE_URL}/api/member/department/queryCompany",
            headers=headers,
            json={"companyId": company_id},
            timeout=15,
        )
        .json()
        .get("result", {})
        .get("users", [])
    )
    if cache is not None:
        cache[cache_key] = users
    return users


def default_fee_payload():
    return {
        "icon": "md-plane",
        "iconColor": "#4c7cc3",
        "feeJson": [
            {
                "disableDel": True,
                "name": "amount",
                "id": 14,
                "label": "报销金额",
                "type": "j-amount",
                "props": {
                    "minValue": 0.00,
                    "maxValue": 999999999.99,
                    "placeholder": "请输入报销金额",
                    "defaultValueType": 1,
                    "title": "报销金额",
                    "required": True,
                },
            },
            {
                "disableDel": False,
                "name": "time",
                "id": 9,
                "label": "日期",
                "type": "j-date",
                "props": {
                    "canEdit": True,
                    "placeholder": "请选择日期",
                    "defaultValueType": 1,
                    "title": "日期",
                    "required": True,
                    "needInputTime": False,
                },
            },
            {
                "name": "consumeCause",
                "id": 39,
                "label": "消费事由",
                "type": "j-text",
                "props": {
                    "min": 0,
                    "max": 64,
                    "defaultValue": "",
                    "canEdit": True,
                    "textType": 1,
                    "placeholder": "请输入消费事由",
                    "title": "消费事由",
                    "required": False,
                },
            },
        ],
        "applyJson": [
            {
                "disableDel": True,
                "name": "amount",
                "id": 14,
                "label": "申请金额",
                "type": "j-amount",
                "props": {
                    "minValue": 0.01,
                    "maxValue": 999999999.99,
                    "placeholder": "请输入申请金额",
                    "defaultValueType": 1,
                    "title": "申请金额",
                    "required": True,
                },
            },
            {
                "disableDel": False,
                "name": "time",
                "id": 9,
                "label": "日期",
                "type": "j-date",
                "props": {
                    "canEdit": True,
                    "placeholder": "请选择日期",
                    "defaultValueType": 1,
                    "title": "日期",
                    "required": True,
                    "needInputTime": False,
                },
            },
            {
                "name": "describe",
                "id": 32,
                "label": "描述",
                "type": "j-text",
                "props": {
                    "min": 0,
                    "max": 14,
                    "defaultValue": "",
                    "canEdit": True,
                    "textType": 1,
                    "placeholder": "请输入描述信息",
                    "title": "描述",
                    "required": False,
                },
            },
        ],
    }


def get_invoice_component(company_id, headers):
    resp = requests.get(
        f"{BASE_URL}/api/bill/component/queryComponentByType",
        headers=headers,
        params={"companyId": company_id, "type": "j-invoice"},
        timeout=12,
    ).json()
    if is_ok(resp) and resp.get("result"):
        return resp["result"].get("props")
    return None


def get_fee_template_detail(fee_id, company_id, headers, detail_cache=None, force_refresh=False):
    cache_key = (company_id, fee_id)
    if detail_cache is not None and not force_refresh and cache_key in detail_cache:
        return detail_cache[cache_key]
    resp = requests.get(
        f"{BASE_URL}/api/bill/feeTemplate/getFeeTemplateById",
        headers=headers,
        params={"id": fee_id, "companyId": company_id},
        timeout=12,
    ).json()
    if is_ok(resp):
        result = resp.get("result")
        if detail_cache is not None and result:
            detail_cache[cache_key] = result
        return result
    return None


def wait_for_fee_template_detail(fee_id, company_id, headers, attempts=4, delay=0.8, detail_cache=None):
    for attempt in range(max(1, attempts)):
        detail = get_fee_template_detail(
            fee_id,
            company_id,
            headers,
            detail_cache=detail_cache,
            force_refresh=(attempt > 0),
        )
        if detail:
            return detail
        if attempt + 1 < attempts:
            time.sleep(delay)
    return None


def build_fee_create_payload(name, parent_id, company_id, headers, template_from_id=None, invoice_component=None, detail_cache=None):
    payload = {"name": name, "parentId": parent_id, "companyId": company_id}
    if template_from_id:
        tmpl = get_fee_template_detail(template_from_id, company_id, headers, detail_cache=detail_cache)
        if tmpl:
            payload["icon"] = tmpl.get("icon") or "md-plane"
            payload["iconColor"] = tmpl.get("iconColor") or "#4c7cc3"
            payload["feeJson"] = tmpl.get("feeJson") or []
            payload["applyJson"] = tmpl.get("applyJson") or []
            return payload

    payload.update(default_fee_payload())
    if invoice_component:
        payload["feeJson"].append(invoice_component)
        payload["applyJson"].append(invoice_component)
    return payload


def get_or_create_fee_template(
    fee_name,
    parent_id,
    company_id,
    headers,
    created_cache=None,
    invoice_component=None,
    template_from_id=None,
    detail_cache=None,
):
    if created_cache is None:
        created_cache = {}

    cache_key = (parent_id, fee_name, template_from_id or 0)
    if cache_key in created_cache:
        return created_cache[cache_key]

    create_payload = build_fee_create_payload(
        fee_name,
        parent_id,
        company_id,
        headers,
        template_from_id=template_from_id,
        invoice_component=invoice_component,
        detail_cache=detail_cache,
    )
    create_resp = requests.post(
        f"{BASE_URL}/api/bill/feeTemplate/addFeeTemplate",
        headers=headers,
        json=create_payload,
        timeout=12,
    ).json()

    if is_ok(create_resp):
        new_id = create_resp.get("result")
        if isinstance(new_id, dict):
            new_id = new_id.get("id") or new_id.get("result")
        if new_id:
            new_id = int(new_id)
            if wait_for_fee_template_detail(new_id, company_id, headers, detail_cache=detail_cache):
                created_cache[cache_key] = new_id
                return new_id

    if "重复" in str(create_resp.get("message", "")):
        fee_tree_retry = requests.get(
            f"{BASE_URL}/api/bill/feeTemplate/queryFeeTemplate",
            headers=headers,
            params={"companyId": company_id, "status": 0, "pageSize": 1000},
            timeout=20,
        ).json().get("result", [])

        def walk(nodes):
            for node in nodes or []:
                if node.get("name") == fee_name and node.get("parentId") == parent_id:
                    return node.get("id")
                found = walk(node.get("children") or [])
                if found:
                    return found
            return None

        existing_id = walk(fee_tree_retry)
        if existing_id and wait_for_fee_template_detail(existing_id, company_id, headers, detail_cache=detail_cache):
            created_cache[cache_key] = existing_id
            return existing_id

    return None


def get_role_tree(company_id, headers, cache=None, force_refresh=False):
    cache_key = f"role_tree:{company_id}"
    if cache is not None and not force_refresh and cache_key in cache:
        return cache[cache_key]
    tree = requests.get(
        f"{BASE_URL}/api/member/role/get/tree",
        headers=headers,
        params={"companyId": company_id},
        timeout=12,
    ).json().get("result", [])
    if cache is not None:
        cache[cache_key] = tree
    return tree


def fee_roles_map(company_id, headers, cache=None):
    tree = get_role_tree(company_id, headers, cache=cache)
    fee_group_id = None
    role_map = {}
    for cat in tree:
        if cat.get("name") == "费用角色组":
            fee_group_id = cat.get("id")
            for rr in cat.get("children", []) or []:
                if rr.get("name") and rr.get("id"):
                    role_map[rr["name"]] = rr["id"]
    return fee_group_id, role_map


def ensure_fee_role_group(company_id, headers, cache=None):
    fee_group_id, _ = fee_roles_map(company_id, headers, cache=cache)
    if fee_group_id:
        return fee_group_id

    create_resp = requests.post(
        f"{BASE_URL}/api/member/role/add/group",
        headers=headers,
        json={"companyId": company_id, "name": "费用角色组"},
        timeout=12,
    ).json()
    if is_ok(create_resp):
        invalidate_cache_entry(cache, f"role_tree:{company_id}")
        fee_group_id, _ = fee_roles_map(company_id, headers, cache=cache)
        return fee_group_id
    return None


def ensure_fee_role(role_name, fee_role_group_id, company_id, headers, cache=None):
    _, role_map = fee_roles_map(company_id, headers, cache=cache)
    if role_name in role_map:
        return role_map[role_name]

    add_resp = requests.post(
        f"{BASE_URL}/api/member/role/add",
        headers=headers,
        json={
            "name": role_name,
            "companyId": company_id,
            "parentId": fee_role_group_id,
            "dataType": "FEE_TYPE",
        },
        timeout=12,
    ).json()
    if is_ok(add_resp) or "存在" in str(add_resp.get("message", "")):
        invalidate_cache_entry(cache, f"role_tree:{company_id}")
        _, role_map = fee_roles_map(company_id, headers, cache=cache)
        return role_map.get(role_name)
    return None


def get_role_detail(role_id, headers):
    resp = requests.get(
        f"{BASE_URL}/api/member/role/get/role",
        headers=headers,
        params={"id": role_id},
        timeout=12,
    ).json()
    if is_ok(resp):
        return resp.get("result") or {}
    return None


def extract_standard_role_relation_ids(role_detail):
    user_ids = []
    department_ids = []
    if not isinstance(role_detail, dict):
        return user_ids, department_ids

    data_type = normalize_text(role_detail.get("dataType"))
    for item in role_detail.get("data") or []:
        if not isinstance(item, dict):
            continue
        if data_type == "DEPARTMENT":
            dep_id = item.get("departmentId") or item.get("id")
            if dep_id:
                department_ids.append(dep_id)
            for user in item.get("users") or []:
                user_id = user.get("userId") or user.get("id")
                if user_id:
                    user_ids.append(user_id)
        else:
            user_id = item.get("userId") or item.get("id")
            if user_id:
                user_ids.append(user_id)

    return unique_list(user_ids), unique_list(department_ids)


def extract_fee_role_relations(role_detail):
    relations = {}
    if not isinstance(role_detail, dict):
        return relations

    for item in role_detail.get("data") or []:
        if not isinstance(item, dict):
            continue
        fee_template_id = item.get("id") or item.get("feeTemplateId")
        if not fee_template_id:
            continue
        user_ids = []
        for user in item.get("users") or []:
            user_id = user.get("userId") or user.get("id")
            if user_id:
                user_ids.append(user_id)
        relations[fee_template_id] = unique_list(user_ids)
    return relations


def clear_fee_role_relations(role_id, headers):
    detail = get_role_detail(role_id, headers)
    if detail is None:
        return False, "读取费用角色详情失败"
    for item in detail.get("data") or []:
        relation_id = item.get("roleFeeTemplateId")
        if not relation_id:
            continue
        resp = requests.delete(
            f"{BASE_URL}/api/member/role/delete/feeTemplate",
            headers=headers,
            params={"id": relation_id},
            timeout=12,
        ).json()
        if not is_ok(resp):
            return False, resp.get("message") or f"删除费用角色关系失败: {relation_id}"
    return True, ""


def normalize_text(v):
    """Normalize various cell values to a clean string.

    Note: some Excel headers/rows may produce a pandas Series (e.g. duplicated
    column labels). In that case, pick the first non-null item.
    """
    if isinstance(v, pd.Series):
        if v.empty:
            return ""
        # Prefer first non-null value
        for item in v.tolist():
            if not pd.isna(item):
                v = item
                break
        else:
            return ""

    if pd.isna(v):
        return ""
    text = str(v).replace("\xa0", " ").strip()
    if text.lower() == "nan":
        return ""
    return text


def role_nodes_map(company_id, headers, cache=None, force_refresh=False):
    tree = get_role_tree(company_id, headers, cache=cache, force_refresh=force_refresh)
    role_map = {}

    def walk(nodes, root_name=None, depth=0):
        for node in nodes or []:
            node_name = normalize_text(node.get("name"))
            current_root = root_name or node_name
            if depth >= 1 and node_name and node.get("id"):
                role_map.setdefault(
                    node_name,
                    {
                        "id": node.get("id"),
                        "name": node_name,
                        "dataType": node.get("dataType"),
                        "parentId": node.get("parentId"),
                        "rootName": current_root,
                    },
                )
            walk(node.get("children") or [], current_root, depth + 1)

    walk(tree)
    return role_map


def standard_role_groups(company_id, headers, cache=None, force_refresh=False):
    tree = get_role_tree(company_id, headers, cache=cache, force_refresh=force_refresh)
    group_map = {}
    for node in tree or []:
        group_name = normalize_text(node.get("name"))
        if group_name and node.get("id"):
            group_map[group_name] = {
                "id": node.get("id"),
                "name": group_name,
                "dataType": node.get("dataType"),
            }
    return group_map


def guess_standard_role_config(role_name, role_groups):
    dept_keywords = ["部门负责人", "部门主管", "部门经理"]
    grade_keywords = ["职员", "员工", "管理者", "总监", "专员", "级"]
    duty_keywords = ["负责人", "经理", "主管", "财务", "出纳", "人事", "行政", "总助", "助理"]

    if any(k in role_name for k in dept_keywords):
        return ("职务", "DEPARTMENT")
    if any(k in role_name for k in grade_keywords):
        return ("职级", "COMPANY")
    if any(k in role_name for k in duty_keywords):
        return ("职务", "COMPANY")

    if "职务" in role_groups:
        return ("职务", "COMPANY")
    if "职级" in role_groups:
        return ("职级", "COMPANY")

    for root_name, root_info in role_groups.items():
        if root_name != "费用角色组":
            return (root_name, "COMPANY" if root_info.get("dataType") != "DEPARTMENT" else "DEPARTMENT")
    return (None, None)


def normalize_role_data_type(value):
    text = normalize_text(value).replace(" ", "").upper()
    mapping = {
        "COMPANY": "COMPANY",
        "普通角色": "COMPANY",
        "公司角色": "COMPANY",
        "普通": "COMPANY",
        "公司": "COMPANY",
        "DEPARTMENT": "DEPARTMENT",
        "部门角色": "DEPARTMENT",
        "部门": "DEPARTMENT",
    }
    return mapping.get(text)


def role_matches_data_type(role_info, desired_data_type):
    if not desired_data_type:
        return True
    return normalize_role_data_type((role_info or {}).get("dataType")) == desired_data_type


def choose_standard_role_parent(role_name, role_groups, desired_data_type=None):
    root_name, guessed_data_type = guess_standard_role_config(role_name, role_groups)
    desired_data_type = desired_data_type or guessed_data_type or "COMPANY"

    if root_name and role_groups.get(root_name):
        return role_groups[root_name], desired_data_type
    if role_groups.get("职务"):
        return role_groups["职务"], desired_data_type

    for root_name, root_info in role_groups.items():
        if root_name == "费用角色组":
            continue
        if desired_data_type and normalize_role_data_type(root_info.get("dataType")) == desired_data_type:
            return root_info, desired_data_type

    for root_name, root_info in role_groups.items():
        if root_name != "费用角色组":
            return root_info, desired_data_type
    return None, desired_data_type


def ensure_standard_role(role_name, company_id, headers, cache=None, data_type=None):
    existing_roles = role_nodes_map(company_id, headers, cache=cache)
    if role_name in existing_roles:
        return existing_roles[role_name]

    role_groups = standard_role_groups(company_id, headers, cache=cache)
    parent, resolved_data_type = choose_standard_role_parent(role_name, role_groups, desired_data_type=data_type)
    if not parent:
        return None

    add_resp = requests.post(
        f"{BASE_URL}/api/member/role/add",
        headers=headers,
        json={
            "name": role_name,
            "companyId": company_id,
            "parentId": parent.get("id"),
            "dataType": resolved_data_type or "COMPANY",
        },
        timeout=12,
    ).json()
    if is_ok(add_resp) or "存在" in str(add_resp.get("message", "")):
        invalidate_cache_entry(cache, f"role_tree:{company_id}")
        return role_nodes_map(company_id, headers, cache=cache).get(role_name)
    return None


def add_role_relation(role_info, user_ids, company_id, headers, department_ids=None):
    payload = {
        "roleId": role_info.get("id"),
        "userIds": sorted(set(user_ids)),
    }
    if role_info.get("dataType") == "DEPARTMENT":
        payload["departmentIds"] = sorted(set(department_ids or []))

    return requests.post(
        f"{BASE_URL}/api/member/role/add/relation",
        headers=headers,
        json=payload,
        timeout=12,
    ).json()


def set_user_departments_exact(user_id, department_ids, company_id, headers):
    payload = {
        "companyId": company_id,
        "users": [user_id],
        "departments": unique_list(department_ids),
    }
    return requests.post(
        f"{BASE_URL}/api/member/department/setUsersDepartments",
        headers=headers,
        json=payload,
        timeout=12,
    ).json()


def query_departments(company_id, headers):
    return (
        requests.get(
            f"{BASE_URL}/api/member/department/queryDepartments",
            headers=headers,
            params={"companyId": company_id},
            timeout=15,
        )
        .json()
        .get("result", [])
        or []
    )


def add_department(title, parent_id, company_id, headers):
    sort_order = 1
    try:
        existing_children = (
            requests.get(
                f"{BASE_URL}/api/member/department/queryDepartmentsByParentId",
                headers=headers,
                params={"companyId": company_id, "parentId": parent_id},
                timeout=12,
            )
            .json()
            .get("result", [])
            or []
        )
        sort_order = len(existing_children) + 1
    except Exception:
        pass

    payload = {
        "companyId": company_id,
        "parentId": parent_id,
        "title": title,
        "name": title,
        "status": "ACTIVE",
        "departmentType": "DEP",
        "companyInvoiceId": 0,
        "defaultFlag": False,
        "nameUpdateFlag": 0,
        "sortOrder": sort_order,
    }
    return requests.post(
        f"{BASE_URL}/api/member/department/add",
        headers=headers,
        json=payload,
        timeout=12,
    ).json()


def build_department_index(nodes):
    id_to_node = {}
    children_by_parent = {}

    def walk(items, fallback_parent_id=-1):
        for item in items or []:
            dep_id = item.get("id")
            title = normalize_text(item.get("title"))
            parent_id = item.get("parentId")
            if parent_id is None:
                parent_id = fallback_parent_id
            if dep_id:
                id_to_node[dep_id] = item
                children_by_parent.setdefault(parent_id, {})
                if title:
                    children_by_parent[parent_id].setdefault(title, item)
            walk(item.get("children") or [], dep_id or parent_id)

    walk(nodes, -1)
    return {"nodes": nodes, "id_to_node": id_to_node, "children_by_parent": children_by_parent}


def remember_department_in_index(dep_index, dep_id, title, parent_id):
    dep_index = dep_index or {"nodes": [], "id_to_node": {}, "children_by_parent": {}}
    node = {"id": dep_id, "title": title, "parentId": parent_id}
    dep_index.setdefault("id_to_node", {})[dep_id] = node
    dep_index.setdefault("children_by_parent", {}).setdefault(parent_id, {})[title] = node
    return dep_index


def department_titles_from_row(row, df1):
    dept_titles = []
    available_headers = {str(col).strip() for col in df1.columns if str(col).strip()}

    # Support both historical and current sheet layouts.
    # old: 一级部门名称 -> 二级部门 -> 三级部门
    # new: 企业名称 -> 一级部门 -> 二级部门
    #
    # Some new workbooks were manually edited and may still carry a duplicated
    # "三级部门" header while the actual semantic shift has already happened.
    # In that case, prefer the new layout as long as "企业名称" is present and
    # the old "一级部门名称" header is absent.
    if "企业名称" in available_headers and "一级部门名称" not in available_headers:
        labels = ["企业名称", "一级部门", "二级部门", "三级部门"]
    elif "一级部门名称" in available_headers or "三级部门" in available_headers:
        labels = ["一级部门名称", "二级部门", "三级部门"]
    else:
        labels = ["企业名称", "一级部门", "二级部门"]

    for label in labels:
        try:
            col = get_col(df1, label)
        except KeyError:
            continue
        title = normalize_text(row.get(col, ""))
        if title and title.lower() != "nan":
            dept_titles.append(title)
    return dept_titles


def candidate_department_suffixes(dept_titles, shortest_first=False):
    if shortest_first:
        indexes = range(len(dept_titles) - 1, -1, -1)
    else:
        indexes = range(len(dept_titles))
    return [dept_titles[i:] for i in indexes if dept_titles[i:]]


def is_top_level_department_title(title, dep_index):
    children_by_parent = (dep_index or {}).get("children_by_parent") or {}
    return title in children_by_parent.get(-1, {}) or title in children_by_parent.get(0, {})


def resolve_department_path_ids(dept_titles, dep_index):
    if not dept_titles:
        return []
    children_by_parent = (dep_index or {}).get("children_by_parent") or {}
    parent_id = -1
    path_ids = []
    for title in dept_titles:
        node = children_by_parent.get(parent_id, {}).get(title)
        if (not node or not node.get("id")) and parent_id == -1 and not path_ids:
            node = children_by_parent.get(0, {}).get(title)
        if not node or not node.get("id"):
            return []
        dep_id = node.get("id")
        path_ids.append(dep_id)
        parent_id = dep_id
    return unique_list(path_ids)


def ensure_department_path_ids(dept_titles, company_id, headers, dep_index=None):
    dep_index = dep_index or build_department_index(query_departments(company_id, headers))

    parent_id = -1
    path_ids = []
    for title in dept_titles:
        children_by_parent = dep_index.get("children_by_parent") or {}
        node = children_by_parent.get(parent_id, {}).get(title)
        if (not node or not node.get("id")) and parent_id == -1 and not path_ids:
            node = children_by_parent.get(0, {}).get(title)
        if not node or not node.get("id"):
            created = add_department(title, parent_id, company_id, headers)
            if not (is_ok(created) or "存在" in str(created.get("message", ""))):
                return [], dep_index, f"创建部门失败: {created.get('message')}"
            created_node = (created.get("result") or {}) if isinstance(created, dict) else {}
            created_id = created_node.get("id")
            if created_id:
                dep_index = remember_department_in_index(dep_index, created_id, title, parent_id)
                node = {"id": created_id, "title": title, "parentId": parent_id}
            else:
                dep_index = build_department_index(query_departments(company_id, headers))
                children_by_parent = dep_index.get("children_by_parent") or {}
                node = children_by_parent.get(parent_id, {}).get(title)
                if not node or not node.get("id"):
                    return [], dep_index, "创建后仍未在部门树中找到"
        dep_id = node.get("id")
        path_ids.append(dep_id)
        parent_id = dep_id

    return unique_list(path_ids), dep_index, ""


def flatten_departments(nodes):
    dep_map = {}

    def walk(items):
        for item in items or []:
            title = normalize_text(item.get("title"))
            dep_id = item.get("id")
            if title and dep_id:
                dep_map[title] = dep_id
            walk(item.get("children") or [])

    walk(nodes)
    return dep_map


def unique_list(values):
    result = []
    seen = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def merge_unique_ids(*groups):
    merged = []
    for group in groups:
        for value in group or []:
            if value in (None, ""):
                continue
            merged.append(value)
    return unique_list(merged)


def build_department_path_cache(df1, users):
    path_cache = {}
    users_by_mobile = {}
    users_by_name = {}
    for user in users or []:
        mobile = normalize_mobile(user.get("mobile") or user.get("userName") or user.get("phone"))
        if mobile and mobile not in users_by_mobile:
            users_by_mobile[mobile] = user
        name = normalize_text(user.get("nickName"))
        if name and name not in users_by_name:
            users_by_name[name] = user

    for _, row in df1.iterrows():
        raw_titles = department_titles_from_row(row, df1)
        if not raw_titles:
            continue
        mobile = normalize_mobile(row.get(get_col(df1, "手机号"), ""))
        name = normalize_text(row.get(get_col(df1, "姓名"), ""))
        user = users_by_mobile.get(mobile) or users_by_name.get(name)
        if not user:
            continue
        title_to_id = {}
        for dep in user.get("departments") or []:
            dep_title = normalize_text(dep.get("title"))
            dep_id = dep.get("id")
            if dep_title and dep_id and dep_title not in title_to_id:
                title_to_id[dep_title] = dep_id
        for suffix in candidate_department_suffixes(raw_titles, shortest_first=False):
            if all(title in title_to_id for title in suffix):
                path_cache[tuple(suffix)] = [title_to_id[title] for title in suffix]
                break
    return path_cache


def ensure_department_ids_from_row(row, df1, company_id, headers, dep_index=None, path_cache=None):
    raw_titles = department_titles_from_row(row, df1)
    if not raw_titles:
        return [], raw_titles, dep_index, "部门缺失或无效"

    for suffix in candidate_department_suffixes(raw_titles, shortest_first=False):
        cached_ids = (path_cache or {}).get(tuple(suffix))
        if cached_ids:
            return unique_list(cached_ids), suffix, dep_index, ""
        resolved_ids = resolve_department_path_ids(suffix, dep_index)
        if resolved_ids:
            if path_cache is not None:
                path_cache[tuple(suffix)] = unique_list(resolved_ids)
            return unique_list(resolved_ids), suffix, dep_index, ""

    for suffix in candidate_department_suffixes(raw_titles, shortest_first=True):
        if not is_top_level_department_title(suffix[0], dep_index):
            continue
        dept_ids, dep_index, err = ensure_department_path_ids(suffix, company_id, headers, dep_index=dep_index)
        if not err and dept_ids:
            if path_cache is not None:
                path_cache[tuple(suffix)] = unique_list(dept_ids)
            return unique_list(dept_ids), suffix, dep_index, ""

    return [], raw_titles, dep_index, "未找到可用的部门路径"


def split_values(v):
    t = normalize_text(v)
    if not t or t.lower() == "nan":
        return []
    for ch in ["，", "、", ";", "；"]:
        t = t.replace(ch, ",")
    return [x.strip() for x in t.split(",") if x.strip()]


def normalize_mobile(v):
    t = re.sub(r"\D", "", "" if pd.isna(v) else str(v))
    return t[:11] if t else ""


def build_sheet_user_aliases(df1, users):
    by_mobile = {}
    for u in users:
        uid = u.get("id")
        if not uid:
            continue
        for key in ["mobile", "userName", "phone"]:
            mobile = normalize_mobile(u.get(key))
            if mobile and mobile not in by_mobile:
                by_mobile[mobile] = u

    alias_map = {}
    for _, row in df1.iterrows():
        sheet_name = normalize_text(row.get(get_col(df1, "姓名"), ""))
        mobile = normalize_mobile(row.get(get_col(df1, "手机号"), ""))
        if not (sheet_name and mobile):
            continue
        matched = by_mobile.get(mobile)
        if matched and matched.get("id"):
            alias_map[sheet_name] = matched.get("id")
    return alias_map


def template_defaults_from_model(bill_type, default_model):
    defaults = {
        "businessType": "PRIVATE",
        "componentJson": [],
        "icon": "md-pricetag",
        "iconColor": "#4c7cc3",
        "payFlag": True,
        "requestScope": False,
    }
    if bill_type in {"EXPENSE", "PAYMENT"}:
        defaults.update(
            {
                "applyRelateFlag": True,
                "applyRelateNecessary": False,
                "feeScopeFlag": False,
            }
        )
    elif bill_type == "LOAN":
        defaults.update(
            {
                "applyRelateFlag": False,
                "refundDateFlag": False,
            }
        )
    elif bill_type == "REQUISITION":
        defaults.update(
            {
                "feeScopeFlag": False,
                "applyContentType": "SUMFEE",
                "lessThanApplyAmount": False,
                "relatOnce": False,
            }
        )

    if not default_model:
        return defaults

    for key in [
        "applyRelateFlag",
        "applyRelateNecessary",
        "applyContentType",
        "businessType",
        "componentJson",
        "feeScopeFlag",
        "icon",
        "iconColor",
        "lessThanApplyAmount",
        "loanRelateFlag",
        "loanRelateNecessary",
        "loanRequestScope",
        "needRepayFlag",
        "payFlag",
        "refundDateFlag",
        "relatOnce",
        "requestScope",
    ]:
        if key in default_model and default_model.get(key) is not None:
            defaults[key] = default_model.get(key)
    return defaults


def default_bill_model_source(default_model):
    return normalize_text((default_model or {}).get("_source")).lower()


def uses_fallback_bill_model(default_model):
    return default_bill_model_source(default_model).startswith("fallback")


_TEMPLATE_UPDATE_ALLOWLIST = {
    "applyContentType",
    "applyRelateFlag",
    "applyRelateNecessary",
    "billRejectMode",
    "businessType",
    "closeNumber",
    "closeType",
    "companyId",
    "componentId",
    "componentJson",
    "departmentIds",
    "expensesVisibleRanges",
    "feeDepartScope",
    "feeIds",
    "feeList",
    "feeRoleIds",
    "feeScopeFlag",
    "feeScopeType",
    "groupId",
    "icon",
    "iconColor",
    "id",
    "lessThanApplyAmount",
    "loanIds",
    "loanRelateFlag",
    "loanRelateNecessary",
    "loanRequestScope",
    "loans",
    "name",
    "nameSpell",
    "needRepayFlag",
    "payFlag",
    "printTemplate",
    "refundDateFlag",
    "relatOnce",
    "requisitionIds",
    "requisitions",
    "requestScope",
    "roleIds",
    "scope",
    "status",
    "supplementApply",
    "tripScopeFlag",
    "type",
    "userIds",
    "userScopeFlag",
    "workFlowId",
}


def query_bill_template(template_id, company_id, headers):
    resp = requests.post(
        f"{BASE_URL}/api/bill/template/queryTemplate",
        headers=headers,
        json={"id": template_id, "companyId": company_id},
        timeout=12,
    ).json()
    if is_ok(resp):
        return resp.get("result") or {}
    return None


def extract_template_scope_ids(template_detail):
    scope = (template_detail or {}).get("scope") or {}
    return {
        "departmentIds": merge_unique_ids(
            [item.get("departmentId") or item.get("id") for item in scope.get("departments") or [] if isinstance(item, dict)]
        ),
        "roleIds": merge_unique_ids(
            [item.get("id") for item in scope.get("roles") or [] if isinstance(item, dict)]
        ),
        "userIds": merge_unique_ids(
            [item.get("userId") or item.get("id") for item in scope.get("users") or [] if isinstance(item, dict)]
        ),
    }


def sanitize_template_for_update(template_payload):
    return {k: template_payload.get(k) for k in _TEMPLATE_UPDATE_ALLOWLIST if k in template_payload}


def update_bill_template(template_payload, headers):
    return requests.post(
        f"{BASE_URL}/api/bill/template/updateTemplate",
        headers=headers,
        json=sanitize_template_for_update(template_payload),
        timeout=15,
    ).json()


def build_template_name_id_map(tree):
    name_to_id = {}

    def walk(nodes):
        for node in nodes or []:
            children = node.get("children") or []
            if children:
                walk(children)

            name = normalize_text(node.get("name") or node.get("title"))
            node_id = node.get("id")
            # Heuristic: templates usually have workflow/type/componentId.
            if name and node_id and (
                node.get("workFlowId")
                or node.get("componentId")
                or node.get("type") in {"EXPENSE", "PAYMENT", "LOAN", "REQUISITION"}
            ):
                name_to_id.setdefault(name, node_id)

    walk(tree)
    return name_to_id


def query_template_tree(company_id, headers, cache=None, force_refresh=False):
    cache_key = f"template_tree:{company_id}"
    if cache is not None and not force_refresh and cache_key in cache:
        return cache[cache_key]
    result = (
        requests.get(
            f"{BASE_URL}/api/bill/template/queryTemplateTree",
            headers=headers,
            params={"companyId": company_id},
            timeout=12,
        ).json().get("result", [])
        or []
    )
    if cache is not None:
        cache[cache_key] = result
    return result


def find_template_id_by_name(doc_name, company_id, headers, retries=4, delay=0.8, cache=None):
    for attempt in range(max(1, retries)):
        name_map = build_template_name_id_map(
            query_template_tree(company_id, headers, cache=cache, force_refresh=(attempt > 0))
        )
        template_id = name_map.get(doc_name)
        if template_id:
            return template_id
        if attempt + 1 < retries:
            time.sleep(delay)
    return None


def verify_template_persisted(doc_name, company_id, headers, cache=None):
    template_id = find_template_id_by_name(doc_name, company_id, headers, cache=cache)
    if not template_id:
        return None
    template = query_bill_template(template_id, company_id, headers)
    if not template:
        return None
    return {
        "templateId": template_id,
        "templateName": template.get("name") or doc_name,
        "message": "页面保存未拿到成功提示，但后台已能读取模板",
    }


def ui_save_bill_template_with_retry(doc_name, company_id, headers, preferred_browser="auto", reload_page=False, attempts=3, cache=None):
    errors = []
    for attempt in range(max(1, attempts)):
        try:
            result = ui_save_bill_template(
                doc_name,
                preferred_browser=preferred_browser,
                reload_page=(reload_page or attempt > 0),
            )
            result["attempt"] = attempt + 1
            return "ok", result, errors
        except Exception as exc:
            errors.append(str(exc))
            if attempt + 1 < attempts:
                time.sleep(1 + attempt)

    persisted = verify_template_persisted(doc_name, company_id, headers, cache=cache)
    if persisted:
        persisted["attempt"] = attempts
        return "warning", persisted, errors
    return "fail", None, errors


def read_sheet_with_header(path: Path, sheet: str, header_key: str):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet]

    merged_values = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = worksheet.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = top_left_value

    rows = []
    for row_idx in range(1, worksheet.max_row + 1):
        row_values = []
        for col_idx in range(1, worksheet.max_column + 1):
            row_values.append(merged_values.get((row_idx, col_idx), worksheet.cell(row_idx, col_idx).value))
        rows.append(row_values)

    header_row = None
    for idx, row in enumerate(rows):
        if any(header_key in str(cell) for cell in row if cell is not None):
            header_row = idx
            break
    if header_row is None:
        raise KeyError(f"未找到表头关键字: {header_key}")

    header = [str(col).strip() if col is not None else "" for col in rows[header_row]]
    data_rows = rows[header_row + 1 :]
    df = pd.DataFrame(data_rows, columns=header)
    return df


def _normalize_label(value):
    value = str(value).strip()
    value = re.sub(r"\d+", "", value)
    value = re.sub(r"[（）()\[\]【】:_：\-\s]", "", value)
    return value


def get_col(df, target):
    """Find a column in df by a human label.

    We accept:
    - exact match
    - substring match
    - normalized match (punctuation/whitespace tolerant)

    Backward-compatibility:
    - If a template drops a leading prefix like "单据" (e.g. "适配人员" instead of "单据适配人员"),
      try again with common relaxed aliases.
    """
    target = str(target).strip()
    norm_target = _normalize_label(target)

    def _search(t, nt):
        for col in df.columns:
            c = str(col).strip()
            if c == t:
                return col
        for col in df.columns:
            c = str(col).strip()
            if t and t in c:
                return col
        for col in df.columns:
            c = str(col).strip()
            if not c:
                continue
            norm_col = _normalize_label(c)
            if norm_col == nt or nt in norm_col or norm_col in nt:
                return col
        return None

    found = _search(target, norm_target)
    if found is not None:
        return found

    # relaxed aliases
    aliases = []
    if target.startswith("单据"):
        aliases.append(target.replace("单据", "", 1).strip())
    if target == "单据适配人员":
        aliases.extend(["适配人员", "适配人员（多人用中文逗号）", "适配人员(多人用中文逗号)"])

    for a in aliases:
        fa = _search(a, _normalize_label(a))
        if fa is not None:
            return fa

    raise KeyError(target)


def get_optional_col(df, target):
    try:
        return get_col(df, target)
    except KeyError:
        return None


def has_meaningful_value(value):
    text = str(value).strip()
    return bool(text) and text.lower() not in {"nan", "none", "null"}


def filter_rows_by_optional_flag(df, flag_label=None, required_labels=None):
    if flag_label:
        try:
            flag_col = get_col(df, flag_label)
        except KeyError:
            flag_col = None
        if flag_col:
            mask = df[flag_col].astype(str).str.strip() == "是"
            return df[mask].copy()

    required_cols = []
    for label in required_labels or []:
        try:
            required_cols.append(get_col(df, label))
        except KeyError:
            continue

    if required_cols:
        mask = df[required_cols].apply(lambda row: any(has_meaningful_value(v) for v in row), axis=1)
        return df[mask].copy()

    mask = df.apply(lambda row: any(has_meaningful_value(v) for v in row), axis=1)
    return df[mask].copy()


def collect_row_role_entries(row, df1):
    role_entries = {}
    explicit_role_col = get_optional_col(df1, "角色名称")
    explicit_type_col = get_optional_col(df1, "角色类型")
    legacy_role_col = get_optional_col(df1, "角色管理")

    explicit_data_type = normalize_role_data_type(row.get(explicit_type_col, "")) if explicit_type_col else None
    if explicit_role_col:
        for role_name in split_values(row.get(explicit_role_col, "")):
            role_name = normalize_text(role_name)
            if role_name:
                role_entries[role_name] = {
                    "name": role_name,
                    "dataType": explicit_data_type,
                    "source": "角色名称",
                }

    if legacy_role_col:
        for role_name in split_values(row.get(legacy_role_col, "")):
            role_name = normalize_text(role_name)
            if role_name and role_name not in role_entries:
                role_entries[role_name] = {
                    "name": role_name,
                    "dataType": None,
                    "source": "角色管理",
                }

    return list(role_entries.values())


def normalize_assignment_display(value):
    return "，".join(split_values(value))


def is_instruction_like_text(value):
    text = normalize_text(value)
    if not text:
        return False
    hints = ["说明", "示例", "仅为示例", "根据企业具体管理流程需要制定"]
    return any(hint in text for hint in hints)


def workflow_name_variants(value):
    text = normalize_text(value)
    if not text:
        return []

    variants = [text]
    stripped = re.sub(r"^\d+[\s._-]*", "", text).strip()
    if stripped and stripped != text:
        variants.append(stripped)

    return unique_list(variants)


def resolve_workflow_template_name(value, template_names=None):
    variants = workflow_name_variants(value)
    if not variants:
        return ""

    template_names = template_names or set()
    for candidate in variants:
        if candidate in template_names:
            return candidate

    if template_names:
        suffix_matches = [name for name in template_names if variants[0].endswith(name)]
        if len(suffix_matches) == 1:
            return suffix_matches[0]

    return variants[0]


def build_workflow_doc_map(workflows, template_names=None):
    workflow_map = {}
    for workflow in workflows or []:
        workflow_id = workflow.get("id")
        raw_name = normalize_text(workflow.get("tpName"))
        if not raw_name or not workflow_id:
            continue

        workflow_map.setdefault(raw_name, workflow_id)
        matched_name = resolve_workflow_template_name(raw_name, template_names)
        if matched_name and matched_name != raw_name:
            workflow_map.setdefault(matched_name, workflow_id)
    return workflow_map


def query_workflows(company_id, headers):
    resp = requests.get(
        f"{BASE_URL}/api/bpm/workflow/queryWorkFlow",
        headers=headers,
        params={"companyId": company_id, "size": 200},
        timeout=15,
    ).json()
    if is_ok(resp):
        return resp.get("result", []) or []
    return []


def save_workflow(workflow_id, workflow_name, workflow_json, company_id, headers):
    payload = {
        "id": workflow_id,
        "tpName": workflow_name,
        "workflowJson": workflow_json,
        "companyId": company_id,
    }
    return requests.post(
        f"{BASE_URL}/api/bpm/workflow/addWorkFlow",
        headers=headers,
        json=payload,
        timeout=15,
    ).json()


def query_permission_tree(company_id, headers):
    resp = requests.get(
        f"{BASE_URL}/api/member/permission/tree",
        headers=headers,
        params={"companyId": company_id},
        timeout=15,
    ).json()
    if is_ok(resp):
        return resp.get("result", []) or []
    return []


def update_permission_targets(permission_group_id, role_ids, user_ids, company_id, headers):
    return requests.post(
        f"{BASE_URL}/api/member/permission/update",
        headers=headers,
        json={
            "permissionGroupId": permission_group_id,
            "roleIds": unique_list(role_ids),
            "userIds": unique_list(user_ids),
            "companyId": company_id,
        },
        timeout=15,
    ).json()


def flatten_permission_rows(nodes):
    rows = []

    def walk(items):
        for item in items or []:
            if not isinstance(item, dict):
                continue
            if item.get("id") and normalize_text(item.get("name")):
                rows.append(item)
            walk(item.get("children") or [])

    walk(nodes)
    return rows


def permission_row_keys(permission_row):
    keys = []
    for value in [permission_row.get("name"), permission_row.get("description")]:
        normalized = normalize_text(value)
        if normalized:
            keys.append(normalized)
    return unique_list(keys)


def build_permission_row_map(permission_rows):
    row_map = {}
    for permission_row in permission_rows:
        for key in permission_row_keys(permission_row):
            row_map.setdefault(key, permission_row)
    return row_map


def extract_permission_actor_ids(permission_row):
    role_ids = merge_unique_ids(
        [item.get("id") or item.get("roleId") for item in permission_row.get("roles") or [] if isinstance(item, dict)]
    )
    user_ids = merge_unique_ids(
        [item.get("userId") or item.get("id") for item in permission_row.get("users") or [] if isinstance(item, dict)]
    )
    return role_ids, user_ids


def get_standard_role_member_ids(role_info, headers, role_detail_cache=None):
    role_id = (role_info or {}).get("id")
    if not role_id:
        return None, None
    if role_detail_cache is not None and role_id in role_detail_cache:
        return role_detail_cache[role_id]

    role_detail = get_role_detail(role_id, headers)
    if role_detail is None:
        return None, None

    result = extract_standard_role_relation_ids(role_detail)
    if role_detail_cache is not None:
        role_detail_cache[role_id] = result
    return result


def resolve_workflow_targets(raw_value, user_map, role_map, headers, role_detail_cache=None):
    tokens = split_values(raw_value)
    direct_user_ids = []
    company_roles = []
    department_roles = []
    unknown_tokens = []

    for token in tokens:
        role_info = role_map.get(token)
        if role_info and normalize_role_data_type(role_info.get("dataType")) == "DEPARTMENT":
            department_roles.append(role_info)
            continue
        if role_info and normalize_role_data_type(role_info.get("dataType")) == "COMPANY":
            company_roles.append(role_info)
            continue
        if token in user_map:
            direct_user_ids.append(user_map[token])
            continue
        unknown_tokens.append(token)

    expanded_user_ids = list(direct_user_ids)
    empty_company_roles = []
    role_member_errors = []
    company_role_member_ids = {}
    for role_info in company_roles:
        member_user_ids, _ = get_standard_role_member_ids(role_info, headers, role_detail_cache=role_detail_cache)
        if member_user_ids is None:
            role_member_errors.append(f"读取普通角色 {role_info.get('name')} 成员失败")
            continue
        if not member_user_ids:
            empty_company_roles.append(role_info.get("name"))
            continue
        company_role_member_ids[role_info.get("id")] = unique_list(member_user_ids)

    selected_role = None
    if department_roles:
        selected_role = department_roles[0]
        # A mixed "普通角色 + 部门角色" node only has one role-match slot in the
        # current workflow JSON, so we keep the department role as the role match
        # and expand ordinary-role members as designated approvers.
        for role_info in company_roles:
            expanded_user_ids = merge_unique_ids(
                expanded_user_ids,
                company_role_member_ids.get(role_info.get("id")) or [],
            )
    elif company_roles:
        selected_role = company_roles[0]
        # If multiple ordinary roles are provided, keep the first one as role match
        # and expand any extras to concrete staff so everyone is still included.
        for role_info in company_roles[1:]:
            expanded_user_ids = merge_unique_ids(
                expanded_user_ids,
                company_role_member_ids.get(role_info.get("id")) or [],
            )

    return {
        "tokens": tokens,
        "displayName": "，".join(tokens),
        "userIds": unique_list(expanded_user_ids),
        "selectedRole": selected_role,
        "departmentRole": department_roles[0] if department_roles else None,
        "departmentRoleNames": [normalize_text(role.get("name")) for role in department_roles],
        "extraDepartmentRoleNames": [normalize_text(role.get("name")) for role in department_roles[1:]],
        "companyRole": company_roles[0] if company_roles else None,
        "companyRoleNames": [normalize_text(role.get("name")) for role in company_roles],
        "unknownTokens": unknown_tokens,
        "emptyCompanyRoles": empty_company_roles,
        "roleMemberErrors": role_member_errors,
    }


def resolve_permission_targets(raw_value, user_map, role_map):
    tokens = split_values(raw_value)
    role_ids = []
    user_ids = []
    role_names = []
    user_names = []
    unknown_tokens = []

    for token in tokens:
        role_info = role_map.get(token)
        if role_info and role_info.get("id"):
            role_ids.append(role_info.get("id"))
            role_names.append(role_info.get("name") or token)
            continue
        user_id = user_map.get(token)
        if user_id:
            user_ids.append(user_id)
            user_names.append(token)
            continue
        unknown_tokens.append(token)

    return {
        "tokens": tokens,
        "roleIds": unique_list(role_ids),
        "userIds": unique_list(user_ids),
        "roleNames": unique_list(role_names),
        "userNames": unique_list(user_names),
        "unknownTokens": unknown_tokens,
    }


def workflow_role_selection():
    return {
        "ROLE_TYPE": "",
        "ROLE_ID": "",
        "TYPE": "ROLE",
        "isSelect": False,
    }


def workflow_designation_selection():
    return {
        "STAFFIDS": [],
        "TYPE": "DESIGNATION",
        "isSelect": False,
    }


def workflow_carbon_copy_block():
    return [
        {
            "SELECTIONS": [workflow_role_selection(), workflow_designation_selection()],
            "SENDTIME": "REJECT",
        }
    ]


def workflow_start_node():
    return {
        "TYPE": "start",
        "CARBON_COPY": workflow_carbon_copy_block(),
        "OTHER_CONFIG": {"allowSubmitterRetract": True},
        "NAME": "开始",
    }


def workflow_approval_node(name):
    return {
        "COUNTERSIGN": {
            "POLICY": "ANY",
            "SELECTIONS": [
                {
                    "ROLE_TYPE": "",
                    "ROLE_ID": "",
                    "TYPE": "ROLE",
                    "isSelect": True,
                },
                {
                    "STAFFIDS": [],
                    "TYPE": "DESIGNATION",
                    "isSelect": False,
                },
            ],
        },
        "CONDITION": {
            "OPERATOR": "NULL",
            "SOURCE": "",
            "ID": "",
            "VALUE": 0,
            "TYPE": "",
        },
        "TYPE": "simple",
        "CARBON_COPY": workflow_carbon_copy_block(),
        "OTHER_CONFIG": {
            "allowApproverModify": False,
            "autoAgreeWhenApproverRepeated": True,
            "notActiveNodeHide": True,
            "autoAgreeWhenApproverSameAsSubmitter": True,
            "supportTransferApproval": False,
            "transferApproval": {
                "enableTransfer": False,
                "enableBeforeSign": False,
                "enableAfterSign": False,
            },
        },
        "NAME": name,
    }


def workflow_end_node():
    return {
        "COUNTERSIGN": {
            "POLICY": "ANY",
            "SELECTIONS": [
                {
                    "ROLE_TYPE": "",
                    "ROLE_ID": "",
                    "TYPE": "ROLE",
                    "isSelect": True,
                },
                {
                    "STAFFIDS": [],
                    "TYPE": "DESIGNATION",
                    "isSelect": False,
                },
            ],
        },
        "TYPE": "cashier",
        "NAME": "出纳",
        "CARBON_COPY": [
            {
                "SENDTIME": "REJECT",
                "SELECTIONS": [workflow_role_selection(), workflow_designation_selection()],
            }
        ],
        "CONDITION": {
            "OPERATOR": "NEQ",
            "SOURCE": "expensesType",
            "VALUE": "REQUISITION",
            "TYPE": "STRING",
        },
        "OTHER_CONFIG": {"cashNodeAutoComplete": True},
    }


def build_workflow_staff_refs(user_ids, user_by_id):
    result = []
    for user_id in unique_list(user_ids):
        user = user_by_id.get(user_id) or {}
        result.append({"ID": user_id, "NAME": normalize_text(user.get("nickName")) or str(user_id)})
    return result


def apply_workflow_target_to_selections(selections, user_ids, role_info, user_by_id):
    for selection in selections:
        if selection.get("TYPE") == "ROLE":
            if role_info:
                selection["ROLE_ID"] = role_info.get("id") or ""
                selection["ROLE_TYPE"] = normalize_role_data_type(role_info.get("dataType")) or ""
                selection["isSelect"] = True
            else:
                selection["ROLE_ID"] = ""
                selection["ROLE_TYPE"] = ""
                selection["isSelect"] = False
        elif selection.get("TYPE") == "DESIGNATION":
            staff_refs = build_workflow_staff_refs(user_ids, user_by_id)
            selection["STAFFIDS"] = staff_refs
            selection["isSelect"] = bool(staff_refs)


def build_workflow_json(workflow_name, approval_specs, copy_spec, user_by_id):
    nodes = [workflow_start_node()]
    for spec in approval_specs:
        node = workflow_approval_node(spec["displayName"])
        apply_workflow_target_to_selections(
            node["COUNTERSIGN"]["SELECTIONS"],
            spec.get("userIds") or [],
            spec.get("selectedRole"),
            user_by_id,
        )
        nodes.append(node)

    end_node = workflow_end_node()
    if copy_spec:
        end_node["CARBON_COPY"][0]["SENDTIME"] = "AGREE"
        apply_workflow_target_to_selections(
            end_node["CARBON_COPY"][0]["SELECTIONS"],
            copy_spec.get("userIds") or [],
            copy_spec.get("selectedRole"),
            user_by_id,
        )

    nodes.append(end_node)
    return {
        "NAME": workflow_name,
        "BASIC": {"billRejectMode": "0"},
        "NODES": nodes,
    }


def normalize_result_id(value):
    if isinstance(value, dict):
        return value.get("id") or value.get("result")
    return value


def main():
    parser = argparse.ArgumentParser(description="导入 Agent1 三表到财税通")
    parser.add_argument("--xlsx", required=True, help="Agent1 生成的三表文件")
    parser.add_argument("--output", default="./agent2_import_report.json", help="导入报告输出路径")
    parser.add_argument("--auto-login", action="store_true", help="登录态失效时自动打开浏览器并登录")
    parser.add_argument("--username", help="财税通登录手机号；不传则优先读取 CST_USERNAME，仍缺失时终端提示输入")
    parser.add_argument("--password", help="财税通登录密码；不传则优先读取 CST_PASSWORD，仍缺失时终端隐藏输入")
    parser.add_argument("--company-id", type=int, help="多企业账号时指定 companyId；也可用环境变量 CST_COMPANY_ID")
    parser.add_argument("--company-name", help="期望进入的集团/公司名称；用于校验和多企业切换")
    parser.add_argument(
        "--browser",
        choices=["auto", "edge", "chrome"],
        default="auto",
        help="优先使用的浏览器",
    )
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    token, company_id, _, browser_name = get_auth(
        auto_login=args.auto_login,
        preferred_browser=args.browser,
        username=args.username,
        password=args.password,
        company_id=args.company_id,
        company_name=args.company_name,
        prompt=args.auto_login,
    )
    print(f"✅ 检测到 {browser_name} 浏览器")
    h = {"x-token": token, "Content-Type": "application/json"}
    api_cache = {}
    fee_detail_cache = {}

    report = {
        "companyId": company_id,
        "xlsx": str(xlsx),
        "preflight": {
            "has_risk": False,
            "missing_primary": [],
            "missing_people": [],
            "doc_mismatch_02_only": [],
            "doc_mismatch_03_only": [],
            "workflow_doc_mismatch_only": [],
            "workflow_doc_missing": [],
        },
        "step1": {"ok": 0, "fail": []},
        "step1_department_sync": {"ok": 0, "fail": []},
        "step1_roles": {"ok": 0, "fail": [], "role_by_name": {}, "created_roles": []},
        "step2": {"relations_ok": 0, "relations_fail": [], "role_by_doc": {}, "leaf_by_doc": {}, "bindings_detail": [], "reset_docs": [], "reset_fail": []},
        "step25": {"count": 0, "workflowId": None, "workflowName": None, "workflowByDoc": {}, "created": [], "updated": [], "fail": []},
        "step3": {
            "ok": 0,
            "fail": [],
            "branch_fee_role": [],
            "branch_leaf_fee": [],
            "branch_skip": [],
            "default_model_ok": [],
            "default_model_fail": [],
            "default_model_retry_ok": [],
            "default_model_retry_fail": [],
            "ui_save_ok": [],
            "ui_save_warn": [],
            "ui_save_fail": [],
        },
        "step4": {"ok": 0, "fail": [], "updated_permissions": []},
    }

    # Base maps
    users = query_company_users(company_id, h, cache=api_cache)
    user_map = {u.get("nickName"): u.get("id") for u in users if u.get("nickName") and u.get("id")}
    deps = query_departments(company_id, h)
    dep_index = build_department_index(deps)
    dep_map = flatten_departments(deps)

    # ===== 数据核对阶段 =====
    print("\n" + "="*50)
    print("📋 第一步：核对Excel数据与系统数据")
    print("="*50)
    
    has_error = False
    
    # 1. 查询系统中所有员工
    print("\n1️⃣ 查询系统中现有员工...")
    existing_users = users
    existing_user_names = {u.get("nickName"): u for u in existing_users if u.get("nickName")}
    df1_check = read_sheet_with_header(xlsx, "01_添加员工", "姓名")
    df1_check = filter_rows_by_optional_flag(df1_check, "是否导入", ["姓名", "手机号"])
    existing_user_aliases = build_sheet_user_aliases(df1_check, existing_users)
    print(f"   系统中共有 {len(existing_user_names)} 名员工")
    
    # 2. 查询系统中所有费用科目
    print("\n2️⃣ 查询系统中现有费用科目...")
    fee_tree = requests.get(f"{BASE_URL}/api/bill/feeTemplate/queryFeeTemplate", headers=h, params={"companyId": company_id, "status": 1, "pageSize": 1000}, timeout=20).json().get("result", [])
    
    existing_primary = {}  # 一级科目: {name: id}
    existing_secondary = {}  # 二级科目: {(parent_id, name): id}
    existing_third = {}  # 三级科目: {(parent_id, name): id}
    
    for p in fee_tree:
        if p.get("parentId") == -1:
            existing_primary[str(p.get("name", "")).strip()] = p.get("id")
        else:
            parent_id = p.get("parentId")
            for c in (p.get("children") or []):
                existing_secondary[(parent_id, str(c.get("name", "")).strip())] = c.get("id")
                # 查找三级
                for t3 in (c.get("children") or []):
                    existing_third[(c.get("id"), str(t3.get("name", "")).strip())] = t3.get("id")
    
    print(f"   一级科目: {len(existing_primary)} 个")
    print(f"   二级科目: {len(existing_secondary)} 个")
    print(f"   三级科目: {len(existing_third)} 个")
    
    # 3. 查询系统中所有单据模板
    print("\n3️⃣ 查询系统中现有单据模板...")
    existing_templates = query_template_tree(company_id, h, cache=api_cache)
    existing_template_names = set()
    for g in existing_templates:
        for t in (g.get("children") or []):
            if t.get("name"):
                existing_template_names.add(t.get("name"))
    print(f"   系统中共有 {len(existing_template_names)} 个单据模板")
    
    # 4. 读取Excel并核对
    print("\n4️⃣ 核对 02_费用科目配置 表...")
    df2_check = read_sheet_with_header(xlsx, "02_费用科目配置", "一级费用科目")
    df2_check = filter_rows_by_optional_flag(
        df2_check,
        "是否执行",
        # 新模板可能不再要求“归属单据名称”列（而是用“费用角色名称”绑定到单据模板）
        ["一级费用科目", "二级费用科目", "三级费用科目", "费用角色名称", "单据适配人员"],
    )
    
    # 核对一级费用科目
    missing_primary = []
    checked_primary = set()
    for _, row in df2_check.iterrows():
        p = str(row.get(get_col(df2_check, "一级费用科目"), "")).strip()
        if p and p.lower() != "nan" and p not in checked_primary and p not in existing_primary:
            missing_primary.append(p)
            checked_primary.add(p)
    
    if missing_primary:
        print(f"\n   ❌ 以下一级费用科目在系统中不存在：")
        for p in missing_primary:
            print(f"      - {p}")
        print(f"\n   系统中存在的一级科目：")
        for name in sorted(existing_primary.keys()):
            print(f"      - {name}")
        has_error = True
        report["preflight"]["missing_primary"] = missing_primary
    else:
        print(f"   ✅ 所有一级费用科目都存在")
    
    # 核对人员
    print("\n5️⃣ 核对单据适配人员...")
    missing_people = []
    checked_people = set()

    for _, row in df2_check.iterrows():
        for person in split_values(row.get(get_col(df2_check, "单据适配人员"), "")):
            if person not in checked_people:
                checked_people.add(person)
                if person not in existing_user_names and person not in existing_user_aliases:
                    missing_people.append(person)
    
    if missing_people:
        print(f"\n   ❌ 以下人员在系统中不存在：")
        for p in missing_people:
            print(f"      - {p}")
        print(f"\n   系统中存在的员工（部分）：")
        for name in sorted(existing_user_names.keys())[:15]:
            print(f"      - {name}")
        has_error = True
        report["preflight"]["missing_people"] = missing_people
    else:
        print(f"   ✅ 所有 {len(checked_people)} 名单据适配人员都存在")

    # 核对第一张表里的角色管理
    print("\n5️⃣ 补充核对角色管理列...")
    role_col_check = get_optional_col(df1_check, "角色管理")
    explicit_role_col_check = get_optional_col(df1_check, "角色名称")
    existing_roles = role_nodes_map(company_id, h, cache=api_cache) if (role_col_check or explicit_role_col_check) else {}
    missing_roles = []
    checked_roles = set()

    if role_col_check or explicit_role_col_check:
        for _, row in df1_check.iterrows():
            for role_entry in collect_row_role_entries(row, df1_check):
                role_name = role_entry["name"]
                if not role_name or role_name in checked_roles:
                    continue
                checked_roles.add(role_name)
                role_info = existing_roles.get(role_name)
                if not role_info:
                    missing_roles.append(role_name)
                    continue
                desired_data_type = role_entry.get("dataType")
                if desired_data_type and not role_matches_data_type(role_info, desired_data_type):
                    print(
                        f"   ⚠️  角色 {role_name} 已存在，但系统类型为 {role_info.get('dataType')}，"
                        f"与表格要求的 {desired_data_type} 不一致"
                    )
                    has_error = True
                    report["step1_roles"]["fail"].append(
                        {
                            "角色": role_name,
                            "message": f"系统角色类型为 {role_info.get('dataType')}，与表格要求的 {desired_data_type} 不一致",
                        }
                    )

        if missing_roles:
            print("\n   ⚠️  以下角色当前在系统中不存在，执行时会自动创建：")
            for role_name in missing_roles:
                print(f"      - {role_name}")
        else:
            print(f"   ✅ 角色管理列中的 {len(checked_roles)} 个角色都存在")
    else:
        print("   ℹ️ 第一张表没有检测到“角色管理/角色名称”列，跳过角色绑定预检")

    # 核对 03_单据表
    print("\n6️⃣ 核对 03_单据表...")
    df3_check = read_sheet_with_header(xlsx, "03_单据表", "单据模板名称")
    df3_check = filter_rows_by_optional_flag(df3_check, "是否创建", ["单据分组（一级目录）", "单据模板名称"])
    
    # 检查单据模板名称是否与02表中的“归属单据名称/费用角色名称”匹配
    col_02_doc = None
    for candidate in ("归属单据名称", "费用角色名称", "费用角色名称（与单据模版名称保持三致）"):
        try:
            col_02_doc = get_col(df2_check, candidate)
            break
        except KeyError:
            continue

    if col_02_doc:
        doc_names_from_02 = set(df2_check[col_02_doc].dropna().unique())
        doc_names_from_03 = set(df3_check[get_col(df3_check, "单据模板名称")].dropna().unique())

        mismatch = doc_names_from_02 - doc_names_from_03
        if mismatch:
            print(f"\n   ⚠️  02表中有但03表中没有的单据名称：")
            for d in mismatch:
                print(f"      - {d}")
            report["preflight"]["doc_mismatch_02_only"] = sorted(mismatch)

        mismatch2 = doc_names_from_03 - doc_names_from_02
        if mismatch2:
            print(f"\n   ⚠️  03表中有但02表中没有的单据名称：")
            for d in mismatch2:
                print(f"      - {d}")
            report["preflight"]["doc_mismatch_03_only"] = sorted(mismatch2)
    else:
        print("   ℹ️ 02表未检测到“归属单据名称/费用角色名称”列，跳过02↔03单据名称一致性核对")

    workflow_name_col = None
    workflow_rows_present = False
    try:
        df_wf_check = read_sheet_with_header(xlsx, "审批流", "一级审批")
        df_wf_check = filter_rows_by_optional_flag(df_wf_check, None, ["审批流名称", "一级审批", "二级审批", "三级审批", "抄送人"])
        workflow_name_col = get_col(df_wf_check, "审批流名称")
        if workflow_name_col:
            df_wf_check = df_wf_check[
                ~df_wf_check[workflow_name_col].apply(is_instruction_like_text)
            ].copy()
        workflow_rows_present = not df_wf_check.empty
    except Exception:
        df_wf_check = None

    if workflow_rows_present:
        doc_names_from_03 = {
            normalize_text(value)
            for value in df3_check[get_col(df3_check, "单据模板名称")].tolist()
            if normalize_text(value)
        }
        workflow_doc_names = {
            resolve_workflow_template_name(value, doc_names_from_03)
            for value in df_wf_check[workflow_name_col].tolist()
            if resolve_workflow_template_name(value, doc_names_from_03)
        }

        workflow_only = workflow_doc_names - doc_names_from_03
        if workflow_only:
            print("\n   ⚠️  审批流表中有但03表中没有的单据名称：")
            for doc_name in sorted(workflow_only):
                print(f"      - {doc_name}")
            report["preflight"]["workflow_doc_mismatch_only"] = sorted(workflow_only)

        missing_workflow = doc_names_from_03 - workflow_doc_names
        if missing_workflow:
            print("\n   ⚠️  03表中有但审批流表中没有的单据名称：")
            for doc_name in sorted(missing_workflow):
                print(f"      - {doc_name}")
            report["preflight"]["workflow_doc_missing"] = sorted(missing_workflow)
    
    report["preflight"]["has_risk"] = has_error
    if has_error:
        print("\n" + "="*50)
        print("⚠️ 数据核对发现风险，将继续导入并在报告中保留这些风险项")
        print("="*50)
    else:
        print("\n" + "="*50)
        print("✅ 数据核对通过！")
        print("="*50)

    # Step1
    df1 = read_sheet_with_header(xlsx, "01_添加员工", "姓名")
    df1 = filter_rows_by_optional_flag(df1, "是否导入", ["姓名", "手机号"])
    role_col = get_optional_col(df1, "角色管理")
    role_name_col = get_optional_col(df1, "角色名称")
    existing_user_by_mobile = {}
    for u in users:
        mobile = normalize_mobile(u.get("mobile") or u.get("userName") or u.get("phone"))
        if mobile and mobile not in existing_user_by_mobile:
            existing_user_by_mobile[mobile] = u
    department_path_cache = build_department_path_cache(df1, users)
    for i, row in df1.iterrows():
        name = normalize_text(row.get(get_col(df1, "姓名"), ""))
        mobile = normalize_mobile(row.get(get_col(df1, "手机号"), ""))
        department_ids, department_titles, dep_index, dep_err = ensure_department_ids_from_row(
            row,
            df1,
            company_id,
            h,
            dep_index=dep_index,
            path_cache=department_path_cache,
        )
        if not (name and mobile):
            report["step1"]["fail"].append({"row": int(i + 1), "reason": "姓名/手机号缺失或无效"})
            continue
        if dep_err or not department_ids:
            report["step1"]["fail"].append(
                {"row": int(i + 1), "reason": dep_err or "部门缺失或无效", "departments": department_titles}
            )
            continue
        payload = {"nickName": name, "mobile": mobile, "departmentIds": department_ids, "companyId": company_id}
        r = requests.post(f"{BASE_URL}/api/member/userInfo/add", headers=h, json=payload, timeout=12).json()
        if r.get("code") == 200 or r.get("success"):
            report["step1"]["ok"] += 1
            # 添加成功后，如果返回了用户ID，更新user_map
            new_uid = r.get("result")
            if new_uid:
                user_map[name] = new_uid
                existing_user_by_mobile[mobile] = {
                    "id": new_uid,
                    "nickName": name,
                    "mobile": mobile,
                    "userName": mobile,
                    "departmentIds": list(department_ids),
                }
        else:
            msg = str(r.get("message", ""))
            if "已" in msg or "存在" in msg:
                existing_user = existing_user_by_mobile.get(mobile)
                if existing_user and existing_user.get("id"):
                    merged_department_ids = merge_unique_ids(existing_user.get("departmentIds"), department_ids)
                    update_payload = {
                        "id": existing_user.get("id"),
                        "nickName": name,
                        "mobile": mobile,
                        "departmentIds": merged_department_ids,
                        "companyId": company_id,
                    }
                    update_resp = requests.post(
                        f"{BASE_URL}/api/member/userInfo/update",
                        headers=h,
                        json=update_payload,
                        timeout=12,
                    ).json()
                    if is_ok(update_resp):
                        existing_user["departmentIds"] = merged_department_ids
                        report["step1"]["ok"] += 1
                    else:
                        report["step1"]["fail"].append(
                            {
                                "row": int(i + 1),
                                "reason": f"员工已存在，但更新部门失败: {update_resp.get('message')}",
                                "departments": department_titles,
                            }
                        )
                else:
                    report["step1"]["ok"] += 1
            else:
                report["step1"]["fail"].append({"row": int(i + 1), "reason": msg})

    # Step1 完成后刷新用户列表，确保能获取到所有员工（包括刚添加的和已存在的）
    users = query_company_users(company_id, h, cache=api_cache, force_refresh=True)
    user_map = {u.get("nickName"): u.get("id") for u in users if u.get("nickName") and u.get("id")}
    user_by_mobile = {}
    user_by_id = {}
    for u in users:
        uid = u.get("id")
        if uid and uid not in user_by_id:
            user_by_id[uid] = u
        mobile = normalize_mobile(u.get("mobile") or u.get("userName") or u.get("phone"))
        if mobile and mobile not in user_by_mobile:
            user_by_mobile[mobile] = u
    for alias_name, alias_uid in build_sheet_user_aliases(df1, users).items():
        user_map.setdefault(alias_name, alias_uid)

    # Refresh departments after potential department creation during Step1.
    deps = query_departments(company_id, h)
    dep_index = build_department_index(deps)
    dep_map = flatten_departments(deps)
    department_path_cache = build_department_path_cache(df1, users)

    for i, row in df1.iterrows():
        name = normalize_text(row.get(get_col(df1, "姓名"), ""))
        mobile = normalize_mobile(row.get(get_col(df1, "手机号"), ""))
        department_ids, department_titles, dep_index, dep_err = ensure_department_ids_from_row(
            row,
            df1,
            company_id,
            h,
            dep_index=dep_index,
            path_cache=department_path_cache,
        )
        if not (name and department_ids):
            continue
        if dep_err:
            report["step1_department_sync"]["fail"].append(
                {"row": int(i + 1), "姓名": name, "departments": department_titles, "message": dep_err}
            )
            continue

        user_id = user_map.get(name)
        if not user_id and mobile and user_by_mobile.get(mobile):
            user_id = user_by_mobile[mobile].get("id")
        if not user_id:
            report["step1_department_sync"]["fail"].append(
                {"row": int(i + 1), "姓名": name, "departments": department_titles, "message": "员工导入后未找到，无法精确同步部门"}
            )
            continue

        current_user = user_by_id.get(user_id) or (user_by_mobile.get(mobile) if mobile else None) or {}
        merged_department_ids = merge_unique_ids(current_user.get("departmentIds"), department_ids)
        sync_resp = set_user_departments_exact(user_id, merged_department_ids, company_id, h)
        if is_ok(sync_resp):
            report["step1_department_sync"]["ok"] += 1
        else:
            report["step1_department_sync"]["fail"].append(
                {
                    "row": int(i + 1),
                    "姓名": name,
                    "departments": department_titles,
                    "departmentIds": department_ids,
                    "message": sync_resp.get("message"),
                }
            )

    if role_col or role_name_col:
        role_bindings = {}
        existing_roles = role_nodes_map(company_id, h, cache=api_cache)
        for i, row in df1.iterrows():
            name = normalize_text(row.get(get_col(df1, "姓名"), ""))
            if not name:
                continue
            user_id = user_map.get(name)
            if not user_id:
                report["step1_roles"]["fail"].append({"row": int(i + 1), "姓名": name, "message": "员工导入后未在系统员工列表中找到"})
                continue
            department_ids, _, dep_index, _ = ensure_department_ids_from_row(
                row,
                df1,
                company_id,
                h,
                dep_index=dep_index,
                path_cache=department_path_cache,
            )

            for role_entry in collect_row_role_entries(row, df1):
                role_name = normalize_text(role_entry.get("name"))
                desired_data_type = role_entry.get("dataType")
                if not role_name:
                    continue
                role_info = existing_roles.get(role_name)
                if role_info and desired_data_type and not role_matches_data_type(role_info, desired_data_type):
                    report["step1_roles"]["fail"].append(
                        {
                            "row": int(i + 1),
                            "姓名": name,
                            "角色": role_name,
                            "message": f"系统角色类型为 {role_info.get('dataType')}，与表格要求的 {desired_data_type} 不一致",
                        }
                    )
                    continue
                if not role_info:
                    role_info = ensure_standard_role(role_name, company_id, h, cache=api_cache, data_type=desired_data_type)
                    if role_info:
                        existing_roles = role_nodes_map(company_id, h, cache=api_cache, force_refresh=True)
                        if role_name not in report["step1_roles"]["created_roles"]:
                            report["step1_roles"]["created_roles"].append(role_name)
                    else:
                        report["step1_roles"]["fail"].append({"row": int(i + 1), "姓名": name, "角色": role_name, "message": "角色不存在且自动创建失败"})
                        continue
                if role_info.get("dataType") not in {"COMPANY", "DEPARTMENT"}:
                    report["step1_roles"]["fail"].append({"row": int(i + 1), "姓名": name, "角色": role_name, "message": f"暂不支持给 {role_info.get('dataType')} 类型角色自动加人"})
                    continue
                if role_info.get("dataType") == "DEPARTMENT" and not department_ids:
                    report["step1_roles"]["fail"].append({"row": int(i + 1), "姓名": name, "角色": role_name, "message": "部门角色缺少有效部门，无法绑定"})
                    continue

                binding = role_bindings.setdefault(
                    role_name,
                    {
                        "role": role_info,
                        "user_ids": set(),
                        "department_ids": set(),
                        "users": [],
                    },
                )
                binding["user_ids"].add(user_id)
                if name not in binding["users"]:
                    binding["users"].append(name)
                for dep_id in department_ids:
                    binding["department_ids"].add(dep_id)

        for role_name, binding in role_bindings.items():
            role_info = binding["role"]
            role_detail = get_role_detail(role_info.get("id"), h)
            if role_detail is None:
                report["step1_roles"]["fail"].append(
                    {
                        "角色": role_name,
                        "users": binding["users"],
                        "message": "读取现有角色成员失败，为避免覆盖旧成员已跳过",
                    }
                )
                continue
            existing_user_ids, existing_department_ids = extract_standard_role_relation_ids(role_detail)
            merged_user_ids = merge_unique_ids(existing_user_ids, binding["user_ids"])
            merged_department_ids = None
            if role_info.get("dataType") == "DEPARTMENT":
                merged_department_ids = merge_unique_ids(existing_department_ids, binding["department_ids"])
            rel = add_role_relation(
                role_info,
                merged_user_ids,
                company_id,
                h,
                department_ids=merged_department_ids,
            )
            if is_ok(rel) or "存在" in str(rel.get("message", "")):
                report["step1_roles"]["ok"] += len(binding["user_ids"])
                report["step1_roles"]["role_by_name"][role_name] = {
                    "roleId": role_info.get("id"),
                    "dataType": role_info.get("dataType"),
                    "users": binding["users"],
                }
            else:
                report["step1_roles"]["fail"].append(
                    {
                        "角色": role_name,
                        "users": binding["users"],
                        "message": rel.get("message"),
                    }
                )

    # Fee templates tree - 获取系统中已有的一级科目，只用于验证一级存在性
    fee_tree = requests.get(f"{BASE_URL}/api/bill/feeTemplate/queryFeeTemplate", headers=h, params={"companyId": company_id, "status": 0, "pageSize": 1000}, timeout=20).json().get("result", [])
    primary = {str(p.get("name", "")).strip(): p for p in fee_tree if p.get("parentId") == -1}
    child = {(p.get("id"), str(c.get("name", "")).strip()): c.get("id") for p in fee_tree for c in (p.get("children") or []) if p.get("id") and c.get("name") and c.get("id")}
    invoice_component = get_invoice_component(company_id, h)

    # 创建二级后的ID缓存，key为(parent_id, name)
    created_level2 = {}

    # Step2
    df2 = read_sheet_with_header(xlsx, "02_费用科目配置", "一级费用科目")
    df2 = filter_rows_by_optional_flag(
        df2,
        "是否执行",
        # 新模板可能使用“费用角色名称”替代“归属单据名称”作为单据/角色绑定键
        ["一级费用科目", "二级费用科目", "三级费用科目", "费用角色名称", "单据适配人员"],
    )
    for c in [get_col(df2, "一级费用科目"), get_col(df2, "二级费用科目")]:
        df2[c] = df2[c].ffill()
    # 处理四级费用科目（如果存在）
    if any("四级费用科目" in str(c) for c in df2.columns):
        for c in [get_col(df2, "一级费用科目"), get_col(df2, "二级费用科目"), get_col(df2, "三级费用科目")]:
            df2[c] = df2[c].ffill()

    # Step2 费用科目处理流程：
    # 1. 一级费用科目必须存在（系统已有）
    # 2. 二级不存在则自动创建
    # 3. 三级不存在则自动创建
    # 4. 四级不存在则自动创建
    # 5. 判断归属单据名称和单据适配人员是否同时存在
    # 6. 如果同时存在：
    #    - 确保费用角色组存在（不存在则创建）
    #    - 以“单据模板名称”创建一个费用角色
    #    - 把该单据对应的所有末级费用科目绑定到这个角色
    #    - 把该单据适配的所有人员绑定到这个角色

    fee_role_group_id = ensure_fee_role_group(company_id, h, cache=api_cache)
    _, fee_roles = fee_roles_map(company_id, h, cache=api_cache)
    has_people = {}
    doc_people_map = {}
    row_role_bindings = []
    seen_row_bindings = set()

    # 兼容列名：归属单据名称 / 费用角色名称
    try:
        doc_col = get_col(df2, "归属单据名称")
    except KeyError:
        doc_col = get_col(df2, "费用角色名称")

    for _, row in df2.iterrows():
        doc = normalize_text(row.get(doc_col, ""))
        people = split_values(row.get(get_col(df2, "单据适配人员"), ""))
        if not (doc and people):
            continue
        has_people[doc] = has_people.get(doc, False) or bool(people)
        if doc not in doc_people_map:
            doc_people_map[doc] = []
        for person_name in people:
            if person_name not in doc_people_map[doc]:
                doc_people_map[doc].append(person_name)

    report["step2"]["people_by_doc"] = doc_people_map

    for _, row in df2.iterrows():
        p = normalize_text(row.get(get_col(df2, "一级费用科目"), ""))
        s = normalize_text(row.get(get_col(df2, "二级费用科目"), ""))
        t3 = normalize_text(row.get(get_col(df2, "三级费用科目"), ""))
        t4 = normalize_text(row.get(get_col(df2, "四级费用科目"), "")) if any("四级费用科目" in str(c) for c in df2.columns) else ""
        doc = normalize_text(row.get(doc_col, ""))
        people = split_values(row.get(get_col(df2, "单据适配人员"), ""))
        if not (p and s):
            continue

        primary_info = primary.get(p)
        if not primary_info:
            report["step2"]["relations_fail"].append({"doc": doc, "一级费用科目": p, "message": f"一级费用科目 '{p}' 在系统中不存在"})
            continue
        pid = primary_info.get("id")

        # 先检查是否已创建过（本次运行缓存）
        cache_key_l2 = (pid, s)
        if cache_key_l2 in created_level2:
            sid = created_level2[cache_key_l2]
        else:
            # 检查系统中是否已存在
            sid = child.get(cache_key_l2)
            if sid:
                pass

        # 二级不存在时自动创建，并补齐默认字段模板
        if not sid and pid:
            sid = get_or_create_fee_template(
                s,
                pid,
                company_id,
                h,
                created_cache=created_level2,
                invoice_component=invoice_component,
                template_from_id=pid,
                detail_cache=fee_detail_cache,
            )
            if not sid:
                report["step2"]["relations_fail"].append({"doc": doc, "二级费用科目": s, "message": f"创建二级费用科目 '{s}' 失败或详情不可读"})
                continue

        if not sid:
            report["step2"]["relations_fail"].append({"doc": doc, "二级费用科目": s, "message": f"无法获取二级费用科目 '{s}' 的ID"})
            continue

        leaf_id = sid
        # 缓存刚创建的科目，避免重复查询
        fee_cache = {}

        # 验证三级费用科目名称（不能是纯数字或空）
        if t3:
            if t3.isdigit() or len(t3) < 2:
                report["step2"]["relations_fail"].append({"doc": doc, "三级费用科目": t3, "message": f"三级费用科目名称 '{t3}' 无效（不能是纯数字或单个字符）"})
                continue
            t3_id = get_or_create_fee_template(
                t3,
                sid,
                company_id,
                h,
                fee_cache,
                invoice_component=invoice_component,
                template_from_id=sid,
                detail_cache=fee_detail_cache,
            )
            if t3_id:
                leaf_id = t3_id
            else:
                report["step2"]["relations_fail"].append({"doc": doc, "三级费用科目": t3, "message": f"创建三级费用科目 '{t3}' 失败", "parent_id": sid})
                continue

        # 四级存在时，在三级下查找或创建（如果三级不存在，则直接在二级下创建四级）
        if t4:
            parent_for_t4 = leaf_id if t3 else sid
            t4_id = get_or_create_fee_template(
                t4,
                parent_for_t4,
                company_id,
                h,
                fee_cache,
                invoice_component=invoice_component,
                template_from_id=parent_for_t4,
                detail_cache=fee_detail_cache,
            )
            if t4_id:
                leaf_id = t4_id
            else:
                report["step2"]["relations_fail"].append({"doc": doc, "四级费用科目": t4, "message": f"创建四级费用科目 '{t4}' 失败", "parent_id": parent_for_t4})
                continue

        if doc:
            report["step2"]["leaf_by_doc"].setdefault(doc, [])
            if leaf_id not in report["step2"]["leaf_by_doc"][doc]:
                report["step2"]["leaf_by_doc"][doc].append(leaf_id)

        if doc and people and has_people.get(doc, False):
            ordered_people = []
            seen_people = set()
            for person_name in people:
                if person_name not in seen_people:
                    seen_people.add(person_name)
                    ordered_people.append(person_name)
            binding_key = (doc, leaf_id, tuple(ordered_people))
            if binding_key not in seen_row_bindings:
                seen_row_bindings.add(binding_key)
                row_role_bindings.append(
                    {
                        "doc": doc,
                        "leaf_id": leaf_id,
                        "people": ordered_people,
                    }
                )

    doc_role_ids = {}
    for doc in {binding["doc"] for binding in row_role_bindings}:
        rid = fee_roles.get(doc)
        if not rid and fee_role_group_id:
            rid = ensure_fee_role(doc, fee_role_group_id, company_id, h, cache=api_cache)
            _, fee_roles = fee_roles_map(company_id, h, cache=api_cache)
        if not rid:
            report["step2"]["relations_fail"].append({"doc": doc, "角色名称": doc, "message": "按单据模板名称创建费用角色失败"})
            continue
        doc_role_ids[doc] = rid

    fee_role_state_cache = {}
    for binding in row_role_bindings:
        doc = binding["doc"]
        leaf_id = binding["leaf_id"]
        people = binding["people"]
        rid = doc_role_ids.get(doc)
        if not rid:
            continue

        user_ids = []
        for person_name in people:
            uid = user_map.get(person_name)
            if not uid:
                report["step2"]["relations_fail"].append({"doc": doc, "人员": person_name, "message": f"人员 '{person_name}' 在系统中不存在"})
                continue
            if uid not in user_ids:
                user_ids.append(uid)

        if not user_ids:
            report["step2"]["relations_fail"].append({"doc": doc, "费用科目ID": leaf_id, "message": "本行单据适配人员为空或都未在系统中找到，跳过该费用科目绑定"})
            continue

        if rid not in fee_role_state_cache:
            role_detail = get_role_detail(rid, h)
            if role_detail is None:
                report["step2"]["relations_fail"].append(
                    {"doc": doc, "roleId": rid, "message": "读取现有费用角色关系失败，为避免覆盖旧配置已跳过"}
                )
                continue
            fee_role_state_cache[rid] = extract_fee_role_relations(role_detail)
        existing_user_ids = fee_role_state_cache[rid].get(leaf_id, [])
        merged_user_ids = merge_unique_ids(existing_user_ids, user_ids)

        update_payload = {
            "roleId": rid,
            "companyId": company_id,
            "feeTemplateIds": [leaf_id],
            "userIds": sorted(merged_user_ids),
        }
        rel = requests.post(
            f"{BASE_URL}/api/member/role/add/relation",
            headers=h,
            json=update_payload,
            timeout=12,
        ).json()

        if is_ok(rel):
            fee_role_state_cache[rid][leaf_id] = sorted(merged_user_ids)
            report["step2"]["relations_ok"] += 1
            report["step2"]["role_by_doc"][doc] = [rid]
            report["step2"]["bindings_detail"].append(
                {
                    "doc": doc,
                    "feeTemplateIds": [leaf_id],
                    "people": people,
                    "userIds": sorted(merged_user_ids),
                    "roleId": rid,
                }
            )
        else:
            report["step2"]["relations_fail"].append(
                {
                    "doc": doc,
                    "角色名称": doc,
                    "费用科目ID": leaf_id,
                    "people": people,
                    "message": rel.get("message"),
                }
            )

    # Step2.5
    df3 = read_sheet_with_header(xlsx, "03_单据表", "单据模板名称")
    df3 = filter_rows_by_optional_flag(df3, "是否创建", ["单据分组（一级目录）", "单据模板名称"])
    doc_template_names = {
        normalize_text(value)
        for value in df3[get_col(df3, "单据模板名称")].tolist()
        if normalize_text(value)
    }

    role_map_for_workflow = role_nodes_map(company_id, h, cache=api_cache, force_refresh=True)
    role_detail_cache = {}
    wfs = query_workflows(company_id, h)
    workflow_by_doc = build_workflow_doc_map(wfs, doc_template_names)
    workflow_docs_defined = set()
    fallback_workflow_id = None
    fallback_workflow_name = None
    for workflow in wfs:
        if "通用审批" in str(workflow.get("tpName", "")):
            fallback_workflow_id = workflow.get("id")
            fallback_workflow_name = workflow.get("tpName")
            break
    if not fallback_workflow_id and wfs:
        fallback_workflow_id = wfs[0].get("id")
        fallback_workflow_name = wfs[0].get("tpName")

    report["step25"]["count"] = len(wfs)
    report["step25"]["workflowId"] = fallback_workflow_id
    report["step25"]["workflowName"] = fallback_workflow_name

    try:
        df_wf = read_sheet_with_header(xlsx, "审批流", "一级审批")
        df_wf = filter_rows_by_optional_flag(df_wf, None, ["审批流名称", "一级审批", "二级审批", "三级审批", "抄送人"])
    except Exception:
        df_wf = None

    if df_wf is not None and not df_wf.empty:
        workflow_name_col = get_col(df_wf, "审批流名称")
        if workflow_name_col:
            df_wf = df_wf[
                ~df_wf[workflow_name_col].apply(is_instruction_like_text)
            ].copy()
        copy_name_col = get_optional_col(df_wf, "抄送人")
        for _, row in df_wf.iterrows():
            doc_name = resolve_workflow_template_name(row.get(workflow_name_col, ""), doc_template_names)
            if not doc_name:
                continue
            workflow_docs_defined.add(doc_name)

            approval_specs = []
            row_has_error = False
            for level_label in ["一级审批", "二级审批", "三级审批"]:
                raw_value = row.get(get_col(df_wf, level_label), "")
                display_name = normalize_assignment_display(raw_value)
                if not display_name:
                    continue

                target_spec = resolve_workflow_targets(
                    raw_value,
                    user_map,
                    role_map_for_workflow,
                    h,
                    role_detail_cache=role_detail_cache,
                )
                row_errors = []
                if target_spec["unknownTokens"]:
                    row_errors.append(f"未识别对象: {'，'.join(target_spec['unknownTokens'])}")
                if target_spec["extraDepartmentRoleNames"]:
                    row_errors.append(f"同一节点不支持多个部门角色: {'，'.join(target_spec['departmentRoleNames'])}")
                if target_spec["emptyCompanyRoles"]:
                    row_errors.append(f"普通角色暂无成员: {'，'.join(target_spec['emptyCompanyRoles'])}")
                if target_spec["roleMemberErrors"]:
                    row_errors.extend(target_spec["roleMemberErrors"])
                if not target_spec["userIds"] and not target_spec["selectedRole"]:
                    row_errors.append("没有解析出可用审批对象")

                if row_errors:
                    report["step25"]["fail"].append(
                        {
                            "doc": doc_name,
                            "node": level_label,
                            "name": display_name,
                            "message": "；".join(row_errors),
                        }
                    )
                    row_has_error = True
                    break

                target_spec["displayName"] = display_name
                approval_specs.append(target_spec)

            if row_has_error:
                continue
            if not approval_specs:
                report["step25"]["fail"].append({"doc": doc_name, "message": "审批流至少需要一个审批节点"})
                continue

            copy_spec = None
            if copy_name_col:
                raw_copy_value = row.get(copy_name_col, "")
                copy_display = normalize_assignment_display(raw_copy_value)
                if copy_display:
                    copy_spec = resolve_workflow_targets(
                        raw_copy_value,
                        user_map,
                        role_map_for_workflow,
                        h,
                        role_detail_cache=role_detail_cache,
                    )
                    copy_errors = []
                    if copy_spec["unknownTokens"]:
                        copy_errors.append(f"未识别对象: {'，'.join(copy_spec['unknownTokens'])}")
                    if copy_spec["extraDepartmentRoleNames"]:
                        copy_errors.append(f"同一抄送设置不支持多个部门角色: {'，'.join(copy_spec['departmentRoleNames'])}")
                    if copy_spec["emptyCompanyRoles"]:
                        copy_errors.append(f"普通角色暂无成员: {'，'.join(copy_spec['emptyCompanyRoles'])}")
                    if copy_spec["roleMemberErrors"]:
                        copy_errors.extend(copy_spec["roleMemberErrors"])
                    if not copy_spec["userIds"] and not copy_spec["selectedRole"]:
                        copy_errors.append("没有解析出可用抄送对象")
                    if copy_errors:
                        report["step25"]["fail"].append(
                            {
                                "doc": doc_name,
                                "node": "抄送人",
                                "name": copy_display,
                                "message": "；".join(copy_errors),
                            }
                        )
                        continue

            existing_workflow_id = workflow_by_doc.get(doc_name)
            workflow_json = build_workflow_json(doc_name, approval_specs, copy_spec, user_by_id)
            workflow_resp = save_workflow(existing_workflow_id, doc_name, workflow_json, company_id, h)
            if is_ok(workflow_resp):
                saved_workflow_id = normalize_result_id(workflow_resp.get("result")) or existing_workflow_id
                if saved_workflow_id:
                    workflow_by_doc[doc_name] = saved_workflow_id
                if existing_workflow_id:
                    report["step25"]["updated"].append({"doc": doc_name, "workflowId": saved_workflow_id or existing_workflow_id})
                else:
                    report["step25"]["created"].append({"doc": doc_name, "workflowId": saved_workflow_id})
            else:
                report["step25"]["fail"].append(
                    {
                        "doc": doc_name,
                        "message": workflow_resp.get("message") or "审批流保存失败",
                    }
                )

        wfs = query_workflows(company_id, h)
        workflow_by_doc.update(build_workflow_doc_map(wfs, doc_template_names))
        if not fallback_workflow_id and wfs:
            fallback_workflow_id = wfs[0].get("id")
            fallback_workflow_name = wfs[0].get("tpName")
            report["step25"]["workflowId"] = fallback_workflow_id
            report["step25"]["workflowName"] = fallback_workflow_name

    report["step25"]["workflowByDoc"] = workflow_by_doc

    # 必须有审批流才能继续
    if not fallback_workflow_id and not workflow_by_doc:
        report["step3"]["fail"].append({"doc": "所有", "message": "系统中没有可用的审批流，且审批流表未成功创建审批流"})
        Path(args.output).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print("❌ 导入失败：没有可用的审批流")
        print(json.dumps({
            "step1_ok": report["step1"]["ok"],
            "step1_roles_ok": report["step1_roles"]["ok"],
            "step2_relations_ok": report["step2"]["relations_ok"],
            "step3_ok": report["step3"]["ok"],
            "output": args.output,
        }, ensure_ascii=False, indent=2))
        return

    # Step3
    roles_vis = {}
    tree_all = get_role_tree(company_id, h, cache=api_cache)
    for cat in tree_all:
        if cat.get("name") == "费用角色组":
            continue
        for rr in cat.get("children", []) or []:
            if rr.get("name") and rr.get("id"):
                roles_vis[rr["name"]] = rr["id"]

    groups = existing_templates
    group_map = {g.get("name") or g.get("title"): g.get("id") for g in groups if (g.get("name") or g.get("title")) and g.get("id")}
    template_name_map = build_template_name_id_map(groups)
    missing_template_names = sorted(name for name in doc_template_names if name not in template_name_map)
    if missing_template_names:
        try:
            page_template_map = ui_template_name_id_map(preferred_browser=args.browser, reload_page=False)
        except Exception as exc:
            report["step3"]["default_model_fail"].append(
                {
                    "type": "template-tree-ui",
                    "message": str(exc),
                }
            )
        else:
            for name in missing_template_names:
                if page_template_map.get(name):
                    template_name_map.setdefault(name, page_template_map[name])

    df3[get_col(df3, "单据分组（一级目录）")] = df3[get_col(df3, "单据分组（一级目录）")].ffill()

    type_map = {"报销单": "EXPENSE", "借款单": "LOAN", "批量付款单": "PAYMENT", "申请单": "REQUISITION"}

    created_docs = []
    default_bill_models = {}
    for _, row in df3.iterrows():
        group_name = str(row.get(get_col(df3, "单据分组（一级目录）"), "")).strip()
        doc_type = str(row.get(get_col(df3, "单据大类（二级目录）"), "")).strip()
        doc_name = str(row.get(get_col(df3, "单据模板名称"), "")).strip()
        vis_type = str(row.get(get_col(df3, "可见范围类型"), "")).strip()
        vis_obj = str(row.get(get_col(df3, "可见范围对象"), "")).strip()
        assigned_workflow_id = workflow_by_doc.get(doc_name)
        if not assigned_workflow_id and doc_name not in workflow_docs_defined:
            assigned_workflow_id = fallback_workflow_id
        if not assigned_workflow_id:
            report["step3"]["fail"].append({"doc": doc_name, "message": "未找到该单据可用的审批流，已跳过模板创建/更新"})
            continue

        if group_name not in group_map:
            create_group_resp = requests.post(
                f"{BASE_URL}/api/bill/template/createTemplateGroup",
                headers=h,
                json={"name": group_name, "companyId": company_id},
                timeout=12,
            ).json()
            group_result = create_group_resp.get("result")
            group_id = group_result.get("id") if isinstance(group_result, dict) else group_result
            invalidate_cache_entry(api_cache, f"template_tree:{company_id}")
            if group_id:
                group_map[group_name] = group_id
            else:
                time.sleep(0.4)
                groups = query_template_tree(company_id, h, cache=api_cache, force_refresh=True)
                group_map = {g.get("name") or g.get("title"): g.get("id") for g in groups if (g.get("name") or g.get("title")) and g.get("id")}
                template_name_map = build_template_name_id_map(groups)

        targets = split_values(vis_obj)
        role_ids = [roles_vis[t] for t in targets if vis_type == "角色" and t in roles_vis]
        user_ids = [user_map[t] for t in targets if vis_type == "员工" and t in user_map]
        dep_ids = [dep_map[t] for t in targets if vis_type == "部门" and t in dep_map]

        # 判断是否有有效的可见范围限制
        # 类型是"限制"且有具体对象时，才限制可见范围
        has_targets = bool(targets) and bool(role_ids or user_ids or dep_ids)
        is_limited_type = vis_type == "限制" or vis_type == "角色" or vis_type == "员工" or vis_type == "部门"
        has_visibility = is_limited_type and has_targets
        bill_type = type_map.get(doc_type, "EXPENSE")

        if bill_type not in default_bill_models:
            try:
                default_bill_models[bill_type] = get_default_bill_model(
                    bill_type,
                    preferred_browser=args.browser,
                    group_id=group_map.get(group_name) or 0,
                    fresh_page=False,
                )
                report["step3"]["default_model_ok"].append(
                    {
                        "type": bill_type,
                        "source": default_bill_models[bill_type].get("_source", "browser"),
                        "component_count": len(default_bill_models[bill_type].get("componentJson") or []),
                    }
                )
            except Exception as exc:
                default_bill_models[bill_type] = {}
                report["step3"]["default_model_fail"].append({"type": bill_type, "message": str(exc)})
        model_defaults = template_defaults_from_model(bill_type, default_bill_models.get(bill_type) or {})

        payload = {
            "applyRelateFlag": model_defaults.get("applyRelateFlag", True),
            "applyRelateNecessary": model_defaults.get("applyRelateNecessary", False),
            "businessType": model_defaults.get("businessType", "PRIVATE"),
            "companyId": company_id,
            "componentJson": model_defaults.get("componentJson") or [],
            "departmentIds": dep_ids if has_visibility else [],
            "feeIds": [],
            "feeScopeFlag": model_defaults.get("feeScopeFlag", False),
            "groupId": group_map.get(group_name),
            "icon": model_defaults.get("icon", "md-pricetag"),
            "iconColor": model_defaults.get("iconColor", "#4c7cc3"),
            "loanIds": [],
            "name": doc_name,
            "payFlag": model_defaults.get("payFlag", True),
            "requestScope": model_defaults.get("requestScope", False),
            "requisitionIds": [],
            "roleIds": role_ids if has_visibility else [],
            "status": "ACTIVE",
            "type": bill_type,
            "userIds": user_ids if has_visibility else [],
            "userScopeFlag": has_visibility,
            "workFlowId": assigned_workflow_id,
        }
        for extra_key in [
            "applyContentType",
            "lessThanApplyAmount",
            "loanRelateFlag",
            "loanRelateNecessary",
            "loanRequestScope",
            "needRepayFlag",
            "refundDateFlag",
            "relatOnce",
        ]:
            if extra_key in model_defaults:
                payload[extra_key] = model_defaults[extra_key]

        # 费用限制分支
        # 只有当单据已成功匹配到费用角色且角色里有人时，才勾选“限制费用类型”
        fee_role_ids = report["step2"]["role_by_doc"].get(doc_name, [])
        if has_people.get(doc_name, False) and fee_role_ids:
            payload["feeRoleIds"] = fee_role_ids
            payload["feeScopeType"] = "FEE_ROLE"
            payload["feeIds"] = []
            payload["feeScopeFlag"] = True
            report["step3"]["branch_fee_role"].append({"doc": doc_name, "feeRoleIds": fee_role_ids})
        else:
            # 没有费用角色人员时，不勾选“限制费用类型”
            report["step3"]["branch_skip"].append(doc_name)

        existing_template_id = template_name_map.get(doc_name)
        if existing_template_id:
            existing = query_bill_template(existing_template_id, company_id, h)
            if not existing:
                report["step3"]["fail"].append({"doc": doc_name, "message": "模板已存在但读取详情失败，无法更新"})
                continue

            existing_scope_ids = extract_template_scope_ids(existing)
            merged_department_ids = merge_unique_ids(existing_scope_ids.get("departmentIds"), dep_ids if has_visibility else [])
            merged_role_ids = merge_unique_ids(existing_scope_ids.get("roleIds"), role_ids if has_visibility else [])
            merged_user_ids = merge_unique_ids(existing_scope_ids.get("userIds"), user_ids if has_visibility else [])

            update_fields = dict(payload)
            # Avoid wiping existing field settings if our defaults are empty for any reason.
            if not update_fields.get("componentJson") and existing.get("componentJson"):
                update_fields.pop("componentJson", None)

            merged = dict(existing)
            for k, v in update_fields.items():
                merged[k] = v
            merged["id"] = existing_template_id
            merged["companyId"] = company_id
            merged["departmentIds"] = merged_department_ids
            merged["roleIds"] = merged_role_ids
            merged["userIds"] = merged_user_ids
            merged["userScopeFlag"] = bool(merged_department_ids or merged_role_ids or merged_user_ids)
            if has_people.get(doc_name, False) and fee_role_ids:
                merged["feeRoleIds"] = merge_unique_ids(existing.get("feeRoleIds"), fee_role_ids)
                merged["feeScopeType"] = "FEE_ROLE"
                merged["feeIds"] = []
                merged["feeScopeFlag"] = True

            ur = update_bill_template(merged, h)
            if not is_ok(ur) and uses_fallback_bill_model(default_bill_models.get(bill_type)):
                try:
                    refreshed_model = get_default_bill_model(
                        bill_type,
                        preferred_browser=args.browser,
                        group_id=group_map.get(group_name) or 0,
                        fresh_page=False,
                    )
                except Exception as exc:
                    report["step3"]["default_model_retry_fail"].append(
                        {"doc": doc_name, "type": bill_type, "message": str(exc)}
                    )
                else:
                    if not uses_fallback_bill_model(refreshed_model):
                        default_bill_models[bill_type] = refreshed_model
                        refreshed_defaults = template_defaults_from_model(bill_type, refreshed_model or {})
                        merged["applyRelateFlag"] = refreshed_defaults.get("applyRelateFlag", merged.get("applyRelateFlag", True))
                        merged["applyRelateNecessary"] = refreshed_defaults.get(
                            "applyRelateNecessary", merged.get("applyRelateNecessary", False)
                        )
                        merged["businessType"] = refreshed_defaults.get("businessType", merged.get("businessType", "PRIVATE"))
                        merged["componentJson"] = refreshed_defaults.get("componentJson") or merged.get("componentJson") or []
                        merged["feeScopeFlag"] = refreshed_defaults.get("feeScopeFlag", merged.get("feeScopeFlag", False))
                        merged["icon"] = refreshed_defaults.get("icon", merged.get("icon", "md-pricetag"))
                        merged["iconColor"] = refreshed_defaults.get("iconColor", merged.get("iconColor", "#4c7cc3"))
                        merged["payFlag"] = refreshed_defaults.get("payFlag", merged.get("payFlag", True))
                        merged["requestScope"] = refreshed_defaults.get("requestScope", merged.get("requestScope", False))
                        for extra_key in [
                            "applyContentType",
                            "lessThanApplyAmount",
                            "loanRelateFlag",
                            "loanRelateNecessary",
                            "loanRequestScope",
                            "needRepayFlag",
                            "refundDateFlag",
                            "relatOnce",
                        ]:
                            if extra_key in refreshed_defaults:
                                merged[extra_key] = refreshed_defaults[extra_key]
                            else:
                                merged.pop(extra_key, None)
                        ur = update_bill_template(merged, h)
                        report["step3"]["default_model_retry_ok"].append(
                            {
                                "doc": doc_name,
                                "type": bill_type,
                                "source": refreshed_model.get("_source", "browser"),
                                "success": is_ok(ur),
                            }
                        )
            if is_ok(ur):
                report["step3"]["ok"] += 1
            else:
                report["step3"]["fail"].append({"doc": doc_name, "message": f"更新失败: {ur.get('message')}"})
            continue

        cr = requests.post(f"{BASE_URL}/api/bill/template/createTemplate", headers=h, json=payload, timeout=15).json()
        if cr.get("code") == 200 and cr.get("success"):
            report["step3"]["ok"] += 1
            created_docs.append(doc_name)
            invalidate_cache_entry(api_cache, f"template_tree:{company_id}")
            # Best-effort refresh template map so a later row can update it.
            if cr.get("result"):
                template_name_map.setdefault(doc_name, cr.get("result"))
        else:
            report["step3"]["fail"].append({"doc": doc_name, "message": cr.get("message")})

    if created_docs:
        print("\n7️⃣ 页面保存闭环...")
        for idx, doc_name in enumerate(created_docs):
            save_status, save_result, save_errors = ui_save_bill_template_with_retry(
                doc_name,
                company_id,
                h,
                preferred_browser=args.browser,
                reload_page=(idx == 0),
                attempts=3,
                cache=api_cache,
            )
            if save_status == "ok":
                report["step3"]["ui_save_ok"].append(
                    {
                        "doc": doc_name,
                        "message": save_result.get("message"),
                        "templateId": save_result.get("templateId"),
                        "attempt": save_result.get("attempt"),
                    }
                )
                print(f"   ✅ 已页面保存：{doc_name}")
                continue
            if save_status == "warning":
                report["step3"]["ui_save_warn"].append(
                    {
                        "doc": doc_name,
                        "message": save_result.get("message"),
                        "templateId": save_result.get("templateId"),
                        "errors": save_errors,
                    }
                )
                print(f"   ⚠️ 页面保存告警：{doc_name} -> {save_result.get('message')}")
                continue

            report["step3"]["ui_save_fail"].append(
                {
                    "doc": doc_name,
                    "message": "；".join(save_errors) if save_errors else "页面保存失败",
                }
            )
            print(f"   ❌ 页面保存失败：{doc_name} -> {'；'.join(save_errors) if save_errors else '未知错误'}")

    # Step4
    try:
        df4 = read_sheet_with_header(xlsx, "权限", "权限名称")
        df4 = filter_rows_by_optional_flag(df4, None, ["权限名称", "员工姓名"])
    except Exception:
        df4 = None

    if df4 is not None and not df4.empty:
        permission_rows = flatten_permission_rows(query_permission_tree(company_id, h))
        permission_map = build_permission_row_map(permission_rows)
        role_map_for_permission = role_nodes_map(company_id, h, cache=api_cache, force_refresh=True)
        permission_name_col = get_col(df4, "权限名称")
        permission_actor_col = get_col(df4, "员工姓名")

        for _, row in df4.iterrows():
            permission_name = normalize_text(row.get(permission_name_col, ""))
            actor_text = normalize_text(row.get(permission_actor_col, ""))
            if not permission_name:
                continue
            if not actor_text:
                continue

            permission_row = permission_map.get(permission_name)
            if not permission_row:
                report["step4"]["fail"].append({"permission": permission_name, "message": "系统中未找到该权限项"})
                continue
            if permission_row.get("canEdit") is False:
                report["step4"]["fail"].append({"permission": permission_name, "message": "该权限项当前不支持编辑"})
                continue

            resolved_targets = resolve_permission_targets(
                actor_text,
                user_map,
                role_map_for_permission,
            )
            if resolved_targets["unknownTokens"]:
                report["step4"]["fail"].append(
                    {
                        "permission": permission_name,
                        "message": f"未识别对象: {'，'.join(resolved_targets['unknownTokens'])}",
                    }
                )
                continue
            if not resolved_targets["roleIds"] and not resolved_targets["userIds"]:
                report["step4"]["fail"].append({"permission": permission_name, "message": "未解析出可添加的角色或员工"})
                continue

            existing_role_ids, existing_user_ids = extract_permission_actor_ids(permission_row)
            merged_role_ids = merge_unique_ids(existing_role_ids, resolved_targets["roleIds"])
            merged_user_ids = merge_unique_ids(existing_user_ids, resolved_targets["userIds"])
            permission_resp = update_permission_targets(
                permission_row.get("id"),
                merged_role_ids,
                merged_user_ids,
                company_id,
                h,
            )
            if is_ok(permission_resp):
                report["step4"]["ok"] += 1
                report["step4"]["updated_permissions"].append(
                    {
                        "permission": permission_name,
                        "roleNames": resolved_targets["roleNames"],
                        "userNames": resolved_targets["userNames"],
                    }
                )
            else:
                report["step4"]["fail"].append(
                    {
                        "permission": permission_name,
                        "message": permission_resp.get("message") or "权限更新失败",
                    }
                )

    Path(args.output).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print("✅ 导入完成")
    print(json.dumps({
        "step1_ok": report["step1"]["ok"],
        "step1_roles_ok": report["step1_roles"]["ok"],
        "step2_relations_ok": report["step2"]["relations_ok"],
        "step3_ok": report["step3"]["ok"],
        "step4_ok": report["step4"]["ok"],
        "output": args.output,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
